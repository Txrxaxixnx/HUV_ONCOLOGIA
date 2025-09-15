#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesador independiente y especializado para informes de Autopsia del HUV.

Objetivo: extraer datos estructurados desde PDFs de AUTOPSIA y generar un Excel
con el mismo esquema/plantilla usado por los otros procesadores (55 columnas),
sin modificar `huv_constants.py` ni el flujo principal.
"""

import re
import sys
import unicodedata
from pathlib import Path
from datetime import datetime, date
from dateutil.relativedelta import relativedelta

# Dependencias de terceros
import pandas as pd
import fitz  # PyMuPDF
from PIL import Image
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tkinter as tk
from tkinter import filedialog

# Módulos del proyecto necesarios (no tocar huv_constants)
try:
    from ocr_processing import pdf_to_text_enhanced
    from huv_constants import HUV_CONFIG, CUPS_CODES, PROCEDIMIENTOS, PATTERNS_HUV
except ImportError:
    print("ERROR: Asegúrese de que 'ocr_processing.py' y 'huv_constants.py' estén accesibles.")
    sys.exit(1)


# =============== Utilidades básicas (locales al procesador) ===============
def _normalize_text(u: str) -> str:
    return unicodedata.normalize('NFKD', u or '').encode('ASCII', 'ignore').decode().upper()


def split_full_name(full_name: str) -> dict:
    if not full_name:
        return {'primer_nombre': '', 'segundo_nombre': '', 'primer_apellido': '', 'segundo_apellido': ''}
    parts = [p.strip() for p in full_name.strip().split() if p.strip()]
    result = {'primer_nombre': '', 'segundo_nombre': '', 'primer_apellido': '', 'segundo_apellido': ''}
    if len(parts) == 1:
        result['primer_nombre'] = parts[0]
    elif len(parts) == 2:
        result['primer_nombre'] = parts[0]
        result['primer_apellido'] = parts[1]
    elif len(parts) == 3:
        result['primer_nombre'] = parts[0]
        result['segundo_nombre'] = parts[1]
        result['primer_apellido'] = parts[2]
    elif len(parts) >= 4:
        result['primer_nombre'] = parts[0]
        result['segundo_nombre'] = parts[1]
        result['primer_apellido'] = parts[2]
        result['segundo_apellido'] = ' '.join(parts[3:])
    return result


def calculate_birth_date(edad_str: str, fecha_referencia_str: str = None) -> str:
    try:
        years_match = re.search(r'(\d+)\s*a[ñn]os', edad_str, re.IGNORECASE)
        months_match = re.search(r'(\d+)\s*mes(es)?', edad_str, re.IGNORECASE)
        days_match = re.search(r'(\d+)\s*d[ií]a(s)?', edad_str, re.IGNORECASE)
        years = int(years_match.group(1)) if years_match else 0
        months = int(months_match.group(1)) if months_match else 0
        days = int(days_match.group(1)) if days_match else 0
        ref_date = date.today()
        if fecha_referencia_str:
            for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
                try:
                    ref_date = datetime.strptime(fecha_referencia_str, fmt).date()
                    break
                except ValueError:
                    continue
        birth_date = ref_date - relativedelta(years=years, months=months, days=days)
        return birth_date.strftime('%d/%m/%Y')
    except Exception:
        return ''


def convert_date_format(date_str: str) -> str:
    if not date_str:
        return ''
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime('%d/%m/%Y')
        except ValueError:
            continue
    return date_str


# ======================= Extracción principal ========================
def extract_autopsia_data(text: str) -> dict:
    """Extrae datos de un informe de AUTOPSIA usando patrones base del HUV.

    Mantiene compatibilidad con el mapeo de 55 columnas usado por los
    otros procesadores. No modifica `huv_constants`.
    """
    data: dict = {}

    # 1) Extracción básica con patrones comunes probados
    for key, pattern in PATTERNS_HUV.items():
        try:
            m = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
            data[key] = re.sub(r'\s+', ' ', m.group(1).strip()) if m else ''
        except Exception:
            data[key] = ''

    # 2) Normalizaciones y lógica de negocio específica
    data['tipo_informe'] = 'AUTOPSIA'

    if data.get('nombre_completo'):
        data.update(split_full_name(data['nombre_completo']))

    if data.get('identificacion_numero'):
        data['identificacion_numero'] = re.sub(r'[^\d]', '', data['identificacion_numero'])

    if data.get('edad'):
        ref_date_str = data.get('fecha_autopsia') or data.get('fecha_ingreso') or data.get('fecha_informe')
        data['fecha_nacimiento'] = calculate_birth_date(data['edad'], ref_date_str)
        m_y = re.search(r'(\d+)', data['edad'])
        data['edad'] = m_y.group(1) if m_y else ''

    # Fechas de ordenamiento: para Autopsia usar fecha_autopsia si está
    if data.get('fecha_autopsia'):
        data['fecha_ordenamiento'] = data['fecha_autopsia']
    else:
        data['fecha_ordenamiento'] = data.get('fecha_ingreso', '')

    # CUPS/Procedimiento
    data['cups_code'] = CUPS_CODES.get('AUTOPSIA', '')
    data['procedimiento'] = PROCEDIMIENTOS.get(data['cups_code'], '')

    # Naturaleza de AUTOPSIA
    data['hospitalizado'] = 'SI'
    data['organo_final'] = 'CUERPO HUMANO COMPLETO'

    num = data.get('numero_peticion', '')
    data['specimens'] = [{'muestra': num, 'organo': data['organo_final']}]

    # Identificador único: usar certificado de defunción si está disponible
    if data.get('certificado_defuncion') and not data.get('identificador_unico'):
        data['identificador_unico'] = data['certificado_defuncion']

    # Malignidad por texto general (macro/micro/diagnóstico)
    macro = data.get('descripcion_macroscopica', '')
    micro = data.get('descripcion_microscopica', '')
    dx = data.get('diagnostico', '')
    text_to_check = _normalize_text(' '.join([macro, micro, dx]))
    data['malignidad'] = 'PRESENTE' if re.search(r'\b(CARCINOMA|CANCER|MALIGN|METASTAS|ADENOCARCINOMA|LINFOMA|SARCOMA|MELANOMA)\b', text_to_check) else 'AUSENTE'

    # Especialidad deducida simple para autopsia (mantener genérico)
    data['especialidad_deducida'] = 'MEDICINA GENERAL'

    return data


# ======================= Mapeo a Excel (55 cols) ========================
def map_to_excel_format(extracted_data: dict) -> list:
    EXCEL_COLUMNS = [
        "N. peticion (0. Numero de biopsia)", "Hospitalizado", "Sede", "EPS", "Servicio",
        "Médico tratante", "Especialidad", "Ubicación", "N. Autorizacion", "Identificador Unico",
        "Datos Clinicos", "Fecha ordenamiento", "Tipo de documento", "N. de identificación",
        "Primer nombre", "Segundo nombre", "Primer apellido", "Segundo apellido", "Fecha de nacimiento",
        "Edad", "Genero", "Número celular", "Direccion de correo electronico", "Direccion de correo electronico 2",
        "Contacto de emergencia", "Departamento", "Teléfono del contacto", "Municipio", "N. muestra",
        "CUPS", "Tipo de examen (4, 12, Metodo de obtención de la muestra, factor de certeza para el diagnóstico)",
        "Procedimiento (11. Tipo de estudio para el diagnóstico)", "Organo (1. Muestra enviada a patología)",
        "Tarifa", "Valor", "Copago", "Descuento", "Fecha de ingreso (2. Fecha de la muestra)",
        "Fecha finalizacion (3. Fecha del informe)", "Usuario finalizacion", "Usuario asignacion micro",
        "Fecha asignacion micro", "Malignidad", "Condicion", "Descripcion macroscopica",
        "Descripcion microscopica (8,9, 10,12,. Invasión linfovascular y perineural, indice mitótico/Ki67, Inmunohistoquímica, tamaño tumoral)",
        "Descripcion Diagnostico (5,6,7 Tipo histológico, subtipo histológico, margenes tumorales)",
        "Diagnostico Principal", "Comentario", "Informe adicional", "Congelaciones /Otros estudios",
        "Liquidos (5 Tipo histologico)", "Citometria de flujo (5 Tipo histologico)", "Hora Desc. macro",
        "Responsable macro"
    ]

    rows = []
    specimens = extracted_data.get('specimens', []) or [{'muestra': extracted_data.get('numero_peticion', ''), 'organo': extracted_data.get('organo_final', '')}]

    for specimen in specimens:
        row = {col: '' for col in EXCEL_COLUMNS}

        row["N. peticion (0. Numero de biopsia)"] = extracted_data.get('numero_peticion', '')
        row["Hospitalizado"] = extracted_data.get('hospitalizado', 'SI')
        row["Sede"] = HUV_CONFIG['sede_default']
        row["EPS"] = extracted_data.get('eps', '')
        row["Servicio"] = extracted_data.get('servicio', '')
        row["Médico tratante"] = extracted_data.get('medico_tratante', '')
        row["Especialidad"] = extracted_data.get('especialidad_deducida', '')
        row["Ubicación"] = extracted_data.get('servicio', '')
        row["N. Autorizacion"] = extracted_data.get('n_autorizacion', '')
        row["Identificador Unico"] = extracted_data.get('identificador_unico', '')
        row["Datos Clinicos"] = 'SI' if len(extracted_data.get('descripcion_macroscopica', '')) > 50 else 'NO'
        row["Fecha ordenamiento"] = convert_date_format(extracted_data.get('fecha_ordenamiento', ''))
        row["Tipo de documento"] = extracted_data.get('tipo_documento', HUV_CONFIG['tipo_documento_default'])
        row["N. de identificación"] = extracted_data.get('identificacion_numero', '')
        row["Primer nombre"] = extracted_data.get('primer_nombre', '')
        row["Segundo nombre"] = extracted_data.get('segundo_nombre', '')
        row["Primer apellido"] = extracted_data.get('primer_apellido', '')
        row["Segundo apellido"] = extracted_data.get('segundo_apellido', '')
        row["Fecha de nacimiento"] = extracted_data.get('fecha_nacimiento', '')
        row["Edad"] = extracted_data.get('edad', '')
        row["Genero"] = extracted_data.get('genero', '')

        row["Departamento"] = HUV_CONFIG['departamento_default']
        row["Municipio"] = HUV_CONFIG['municipio_default']
        row["N. muestra"] = specimen.get('muestra', '')
        row["CUPS"] = CUPS_CODES.get('AUTOPSIA', '')
        row["Tipo de examen (4, 12, Metodo de obtención de la muestra, factor de certeza para el diagnóstico)"] = "AUTOPSIA"
        row["Procedimiento (11. Tipo de estudio para el diagnóstico)"] = PROCEDIMIENTOS.get(row["CUPS"], '')
        row["Organo (1. Muestra enviada a patología)"] = specimen.get('organo', extracted_data.get('organo_final', ''))
        row["Tarifa"] = HUV_CONFIG['tarifa_default']
        row["Valor"] = HUV_CONFIG['valor_default']

        row["Fecha de ingreso (2. Fecha de la muestra)"] = convert_date_format(extracted_data.get('fecha_ingreso', ''))
        row["Fecha finalizacion (3. Fecha del informe)"] = convert_date_format(extracted_data.get('fecha_informe', ''))

        # Responsables si existen en el texto OCR
        row["Usuario finalizacion"] = extracted_data.get('usuario_finalizacion', '')
        row["Responsable macro"] = extracted_data.get('responsable_analisis', '')

        row["Malignidad"] = extracted_data.get('malignidad', '')
        row["Descripcion macroscopica"] = extracted_data.get('descripcion_macroscopica', '')
        row["Descripcion microscopica (8,9, 10,12,. Invasión linfovascular y perineural, indice mitótico/Ki67, Inmunohistoquímica, tamaño tumoral)"] = extracted_data.get('descripcion_microscopica', '')
        row["Descripcion Diagnostico (5,6,7 Tipo histológico, subtipo histológico, margenes tumorales)"] = extracted_data.get('diagnostico', '')
        row["Comentario"] = extracted_data.get('comentarios', '')
        row["Diagnostico Principal"] = ''
        row["Hora Desc. macro"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        rows.append({col: row.get(col, '') for col in EXCEL_COLUMNS})

    return rows


# ======================= Ejecución CLI (opcional) ========================
def main():
    print("=" * 60)
    print("PROCESADOR ESPECIALIZADO DE INFORMES DE AUTOPSIA")
    print("=" * 60)

    root = tk.Tk()
    root.withdraw()
    pdf_path = filedialog.askopenfilename(
        title="Seleccione el informe PDF de Autopsia para procesar",
        filetypes=[("Archivos PDF", "*.pdf")]
    )
    if not pdf_path:
        print("No se seleccionó ningún archivo. Saliendo.")
        return

    output_dir = Path(pdf_path).parent

    try:
        print("▶️ Extrayendo texto con OCR...")
        pdf_text = pdf_to_text_enhanced(pdf_path)

        # Artefacto de depuración
        debug_file_path = output_dir / "DEBUG_OCR_OUTPUT_AUTOPSIA.txt"
        with open(debug_file_path, "w", encoding="utf-8") as f:
            f.write(pdf_text)
        print(f"📝 Texto OCR guardado en: {debug_file_path}")

        print("🔎 Aplicando lógica de extracción para Autopsia...")
        extracted = extract_autopsia_data(pdf_text)

        print("📋 Mapeando datos a formato Excel...")
        rows = map_to_excel_format(extracted)
        if not rows:
            print("⚠️ No se obtuvo información mapeable.")
            return

        df = pd.DataFrame(rows)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = output_dir / f"Informe_Autopsia_{timestamp}.xlsx"
        df.to_excel(output_path, index=False, engine="openpyxl")

        # Formato de encabezados
        wb = load_workbook(output_path)
        ws = wb.active
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
        for col in ws.columns:
            try:
                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                ws.column_dimensions[col[0].column_letter].width = min((max_length + 2), 60)
            except ValueError:
                pass
        wb.save(output_path)

        print("\n" + "=" * 60)
        print("✅ PROCESAMIENTO COMPLETADO")
        print(f"📦 Registros generados: {len(rows)}")
        print(f"💾 Archivo: {output_path}")
        print("=" * 60)

    except Exception as e:
        print("\n" + "!" * 60)
        print("❌ Error inesperado durante el procesamiento de Autopsia:")
        print(f"   {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        print("!" * 60)


if __name__ == "__main__":
    main()

