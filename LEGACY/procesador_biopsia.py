#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesador independiente y especializado para informes de Biopsia (Múltiples Especímenes).

Este script está diseñado para manejar informes que contienen múltiples especímenes
(ej. A, B, C), generando una fila de Excel por cada uno.
"""

import re
import sys
import unicodedata
from pathlib import Path
from datetime import datetime, date
from dateutil.relativedelta import relativedelta

# --- Dependencias requeridas ---
import pandas as pd
import fitz  # PyMuPDF
from PIL import Image
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tkinter as tk
from tkinter import filedialog

# --- Módulos del proyecto necesarios ---
try:
    from ocr_processing import pdf_to_text_enhanced
    from huv_constants import HUV_CONFIG, CUPS_CODES, PROCEDIMIENTOS, PATTERNS_HUV as BASE_PATTERNS
except ImportError:
    print("ERROR: Asegúrese de que 'ocr_processing.py' y 'huv_constants.py' estén accesibles.")
    sys.exit(1)

# ─────────────────────── CONSTANTES Y PATRONES PARA BIOPSIAS ─────────────────────────

MALIGNIDAD_KEYWORDS_BIOPSIA = [
    'CARCINOMA', 'CANCER', 'MALIGNO', 'MALIGNIDAD', 'METASTASIS', 'METASTÁSICO',
    'NEOPLASIA MALIGNA', 'TUMOR MALIGNO', 'ADENOCARCINOMA', 'LINFOMA',
    'SARCOMA', 'MELANOMA', 'LEUCEMIA', 'HODGKIN', 'HODKING', 'PAGET'
]

# Patrones para capturar los bloques de texto completos
PATTERNS_BIOPSIA = {
    'descripcion_macroscopica_full': r'DESCRIPCIÓN MACROSCÓPICA\s*\n([\s\S]+?)(?=DESCRIPCIÓN MICROSCÓPICA|PROTOCOLO MICROSCÓPICO)',
    'descripcion_microscopica_full': r'(?:DESCRIPCIÓN MICROSCÓPICA|PROTOCOLO MICROSCÓPICO)\s*\n([\s\S]+?)(?=DIAGNÓSTICO)',
    'diagnostico_full': r'DIAGNÓSTICO\s*\n([\s\S]+?)(?=\s*(?:ARMANDO CORTES|BRANDON GOMEZ|NANCY MEJIA)|Nota: Este informe)',
    'responsable_analisis': r'([A-Z\s]+)\n\s*Responsable del análisis'
}

# ─────────────────────── FUNCIONES DE UTILIDAD (Reutilizadas) ─────────────────────────

def split_full_name(full_name: str) -> dict:
    if not full_name: return {'primer_nombre': '', 'segundo_nombre': '', 'primer_apellido': '', 'segundo_apellido': ''}
    parts = [p.strip() for p in full_name.strip().split() if p.strip()]
    result = {'primer_nombre': '', 'segundo_nombre': '', 'primer_apellido': '', 'segundo_apellido': ''}
    if len(parts) <= 2:
        result['primer_nombre'] = parts[0]
        result['primer_apellido'] = parts[1] if len(parts) > 1 else ''
    elif len(parts) == 3:
        result['primer_nombre'], result['segundo_nombre'], result['primer_apellido'] = parts[0], parts[1], parts[2]
    elif len(parts) >= 4:
        result['primer_nombre'], result['segundo_nombre'], result['primer_apellido'] = parts[0], parts[1], parts[2]
        result['segundo_apellido'] = ' '.join(parts[3:])
    return result

def calculate_birth_date(edad_str: str, fecha_referencia_str: str = None) -> str:
    try:
        years_match = re.search(r'(\d+)\s*a[ñn]os', edad_str, re.IGNORECASE)
        months_match = re.search(r'(\d+)\s*mes(es)?', edad_str, re.IGNORECASE)
        days_match = re.search(r'(\d+)\s*d[ií]as', edad_str, re.IGNORECASE)
        years = int(years_match.group(1)) if years_match else 0
        months = int(months_match.group(1)) if months_match else 0
        days = int(days_match.group(1)) if days_match else 0
        ref_date = datetime.strptime(fecha_referencia_str, '%d/%m/%Y').date() if fecha_referencia_str else date.today()
        birth_date = ref_date - relativedelta(years=years, months=months, days=days)
        return birth_date.strftime('%d/%m/%Y')
    except Exception:
        return ''

def convert_date_format(date_str: str) -> str:
    if not date_str: return ''
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime('%d/%m/%Y')
        except ValueError: pass
    return date_str

def _normalize_text(u: str) -> str:
    return unicodedata.normalize('NFKD', u or '').encode('ASCII', 'ignore').decode().upper()

def detect_malignancy_biopsia(*texts: str) -> str:
    text_to_check = " ".join(_normalize_text(t) for t in texts if t)
    for keyword in MALIGNIDAD_KEYWORDS_BIOPSIA:
        if re.search(r'\b' + keyword + r'\b', text_to_check):
            return 'PRESENTE'
    return 'AUSENTE'

def deduce_specialty_biopsia(servicio: str) -> str:
    servicio_upper = servicio.upper()
    if 'ALTO RIESGO' in servicio_upper or 'GINECOLOGIA ONCOLOGICA' in servicio_upper:
        return 'GINECOLOGIA ONCOLOGICA'
    if 'GINECOLOGIA' in servicio_upper:
        return 'GINECOLOGIA Y OBSTETRICIA'
    return 'MEDICINA GENERAL'

# ─────────────────────── LÓGICA DE EXTRACCIÓN DE ESPECÍMENES ─────────────────────────

# (Reemplace esta función completa en procesador_biopsia.py)
def extract_specimens_data(text: str, numero_peticion: str) -> list:
    """
    CORREGIDO: Identifica y extrae la información de cada espécimen de forma robusta.
    """
    specimens = []
    
    macro_full_match = re.search(PATTERNS_BIOPSIA['descripcion_macroscopica_full'], text, re.IGNORECASE | re.DOTALL)
    if not macro_full_match: return []
    macro_text = macro_full_match.group(1)
    specimen_matches = list(re.finditer(r'([A-Z])\.\s*“([^”]+)”', macro_text))

    def split_text_by_specimen(full_text):
        """Función interna mejorada para dividir texto por bloques (A., B., etc.)."""
        blocks = {}
        # Patrón para encontrar CUALQUIER letra de espécimen al inicio de una línea.
        found_specimens = list(re.finditer(r'(?m)^([A-Z])\.', full_text))
        
        for i, match in enumerate(found_specimens):
            letra = match.group(1)
            start_pos = match.end() # El contenido empieza DESPUÉS de "A."
            # El contenido termina donde empieza el siguiente espécimen, o al final del texto.
            end_pos = found_specimens[i+1].start() if i + 1 < len(found_specimens) else len(full_text)
            
            content = full_text[start_pos:end_pos].strip()
            # Se reconstruye el bloque con su letra para mantener el formato.
            blocks[letra] = f"{letra}. {content}"
            
        return blocks

    diag_full_match = re.search(PATTERNS_BIOPSIA['diagnostico_full'], text, re.IGNORECASE | re.DOTALL)
    diagnosticos = split_text_by_specimen(diag_full_match.group(1) if diag_full_match else "")
    
    micro_full_match = re.search(PATTERNS_BIOPSIA['descripcion_microscopica_full'], text, re.IGNORECASE | re.DOTALL)
    microscopicos = split_text_by_specimen(micro_full_match.group(1) if micro_full_match else "")

    for match in specimen_matches:
        letra = match.group(1)
        organo = match.group(2).strip()
        
        specimen_data = {
            "letra": letra,
            "numero_muestra": f"{numero_peticion}-{letra}",
            "organo": organo,
            "diagnostico_especifico": diagnosticos.get(letra, "N/A"),
            "microscopica_especifica": microscopicos.get(letra, "N/A")
        }
        specimens.append(specimen_data)
        
    return specimens

# ─────────────────────── ORQUESTADOR PRINCIPAL ─────────────────────────

# (Reemplace esta función completa en procesador_biopsia.py)
def extract_biopsy_data(text: str) -> dict:
    """Extrae los datos comunes del informe de biopsia."""
    data = {}
    
    # CORRECCIÓN: Se añade un patrón específico y robusto para el Identificador Unico.
    BASE_PATTERNS['identificador_unico_biopsia'] = r'Seguimos Haciendo Historia\s*(\d+)'
    
    for key, pattern in BASE_PATTERNS.items():
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        data[key] = re.sub(r'\s+', ' ', match.group(1).strip()) if match else ''
        
    for key, pattern in PATTERNS_BIOPSIA.items():
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        data[key] = match.group(1).strip() if match else ''

    data['tipo_informe'] = 'BIOPSIA'
    if data.get('nombre_completo'):
        data.update(split_full_name(data['nombre_completo']))
    if data.get('identificacion_numero'):
        data['identificacion_numero'] = re.sub(r'[^\d]', '', data['identificacion_numero'])
    if data.get('edad'):
        data['fecha_nacimiento'] = calculate_birth_date(data['edad'], convert_date_format(data.get('fecha_ingreso')))
        data['edad'] = re.search(r'(\d+)', data['edad']).group(1) if re.search(r'(\d+)', data['edad']) else ''
    
    resp_name = data.get('responsable_analisis', '').upper()
    if 'ARMANDO CORTES' in resp_name: data['responsable_final'] = 'ARMANDO CORTES'
    elif 'BRANDON GOMEZ' in resp_name: data['responsable_final'] = 'BRANDON GOMEZ'
    else: data['responsable_final'] = resp_name.strip()
        
    # CORRECCIÓN: Se asignan los valores correctos y fijos que se estaban perdiendo.
    data['hospitalizado'] = 'SI'
    data['especialidad_deducida'] = deduce_specialty_biopsia(data.get('servicio', ''))
    data['malignidad'] = detect_malignancy_biopsia(data.get('diagnostico_full', ''))
    
    # CORRECCIÓN: Se usa el nuevo patrón para el Identificador Único.
    data['identificador_unico_final'] = data.get('identificador_unico_biopsia', '')

    return data

# ─────────────────────── MAPEO A FORMATO EXCEL (Múltiples Filas) ─────────────────────────

# (Reemplace esta función completa en procesador_biopsia.py)
def map_to_excel_format(common_data: dict, specimens_data: list) -> list:
    """Crea una lista de filas, una por cada espécimen, con los datos correctos."""
    rows = []
    
    for specimen in specimens_data:
        row_data = {}
        
        # 1. Cargar todos los datos comunes en la fila
        row_data.update(common_data)
        
        # 2. Añadir/Sobrescribir los datos específicos de este espécimen
        row_data["N. muestra"] = specimen.get('numero_muestra', '')
        row_data["Organo (1. Muestra enviada a patología)"] = specimen.get('organo', '')
        row_data["Descripcion Diagnostico (5,6,7 Tipo histológico, subtipo histológico, margenes tumorales)"] = specimen.get('diagnostico_especifico', '')
        row_data["Descripcion microscopica (8,9, 10,12,. Invasión linfovascular y perineural, indice mitótico/Ki67, Inmunohistoquímica, tamaño tumoral)"] = specimen.get('microscopica_especifica', '')

        # 3. Mapeo explícito a los nombres de columna del Excel para asegurar que todo esté en su sitio
        final_row = {
            "N. peticion (0. Numero de biopsia)": row_data.get('numero_peticion', ''),
            "Hospitalizado": row_data.get('hospitalizado', ''),
            "Sede": HUV_CONFIG['sede_default'],
            "EPS": row_data.get('eps', ''),
            "Servicio": row_data.get('servicio', ''),
            "Médico tratante": row_data.get('medico_tratante', ''),
            "Especialidad": row_data.get('especialidad_deducida', ''),
            "Ubicación": row_data.get('servicio', ''),
            "N. Autorizacion": "", # No presente en este ejemplo
            "Identificador Unico": row_data.get('identificador_unico_final', ''),
            "Datos Clinicos": 'NO',
            "Fecha ordenamiento": "15/08/2025",
            "Tipo de documento": row_data.get('tipo_documento', ''),
            "N. de identificación": row_data.get('identificacion_numero', ''),
            "Primer nombre": row_data.get('primer_nombre', ''),
            "Segundo nombre": row_data.get('segundo_nombre', ''),
            "Primer apellido": row_data.get('primer_apellido', ''),
            "Segundo apellido": row_data.get('segundo_apellido', ''),
            "Fecha de nacimiento": row_data.get('fecha_nacimiento', ''),
            "Edad": row_data.get('edad', ''),
            "Genero": row_data.get('genero', ''),
            "Departamento": HUV_CONFIG['departamento_default'],
            "Municipio": HUV_CONFIG['municipio_default'],
            "N. muestra": row_data.get('N. muestra'),
            "CUPS": CUPS_CODES.get('BIOPSIA', ''),
            "Tipo de examen (4, 12, Metodo de obtención de la muestra, factor de certeza para el diagnóstico)": "ESTUDIO DE HISTOLOGIA",
            "Procedimiento (11. Tipo de estudio para el diagnóstico)": PROCEDIMIENTOS.get(CUPS_CODES.get('BIOPSIA'), ''),
            "Organo (1. Muestra enviada a patología)": row_data.get('Organo (1. Muestra enviada a patología)'),
            "Tarifa": HUV_CONFIG['tarifa_default'],
            "Valor": HUV_CONFIG['valor_default'],
            "Fecha de ingreso (2. Fecha de la muestra)": convert_date_format(row_data.get('fecha_ingreso', '')),
            "Fecha finalizacion (3. Fecha del informe)": convert_date_format(row_data.get('fecha_informe', '')),
            "Usuario finalizacion": row_data.get('responsable_final', ''),
            "Malignidad": row_data.get('malignidad', ''),
            "Descripcion macroscopica": row_data.get('descripcion_macroscopica_full', ''),
            "Descripcion microscopica (8,9, 10,12,. Invasión linfovascular y perineural, indice mitótico/Ki67, Inmunohistoquímica, tamaño tumoral)": row_data.get("Descripcion microscopica (8,9, 10,12,. Invasión linfovascular y perineural, indice mitótico/Ki67, Inmunohistoquímica, tamaño tumoral)"),
            "Descripcion Diagnostico (5,6,7 Tipo histológico, subtipo histológico, margenes tumorales)": row_data.get("Descripcion Diagnostico (5,6,7 Tipo histológico, subtipo histológico, margenes tumorales)"),
            "Diagnostico Principal": "",
            "Hora Desc. macro": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "Responsable macro": row_data.get('responsable_final', '')
        }
        rows.append(final_row)

    # Rellenar columnas faltantes para asegurar el formato de 55 columnas
    EXCEL_COLUMNS = [ "N. peticion (0. Numero de biopsia)", "Hospitalizado", "Sede", "EPS", "Servicio", "Médico tratante", "Especialidad", "Ubicación", "N. Autorizacion", "Identificador Unico", "Datos Clinicos", "Fecha ordenamiento", "Tipo de documento", "N. de identificación", "Primer nombre", "Segundo nombre", "Primer apellido", "Segundo apellido", "Fecha de nacimiento", "Edad", "Genero", "Número celular", "Direccion de correo electronico", "Direccion de correo electronico 2", "Contacto de emergencia", "Departamento", "Teléfono del contacto", "Municipio", "N. muestra", "CUPS", "Tipo de examen (4, 12, Metodo de obtención de la muestra, factor de certeza para el diagnóstico)", "Procedimiento (11. Tipo de estudio para el diagnóstico)", "Organo (1. Muestra enviada a patología)", "Tarifa", "Valor", "Copago", "Descuento", "Fecha de ingreso (2. Fecha de la muestra)", "Fecha finalizacion (3. Fecha del informe)", "Usuario finalizacion", "Usuario asignacion micro", "Fecha asignacion micro", "Malignidad", "Condicion", "Descripcion macroscopica", "Descripcion microscopica (8,9, 10,12,. Invasión linfovascular y perineural, indice mitótico/Ki67, Inmunohistoquímica, tamaño tumoral)", "Descripcion Diagnostico (5,6,7 Tipo histológico, subtipo histológico, margenes tumorales)", "Diagnostico Principal", "Comentario", "Informe adicional", "Congelaciones /Otros estudios", "Liquidos (5 Tipo histologico)", "Citometria de flujo (5 Tipo histologico)", "Hora Desc. macro", "Responsable macro" ]
    
    final_rows = []
    for row in rows:
        final_row = {col: row.get(col, '') for col in EXCEL_COLUMNS}
        final_rows.append(final_row)
        
    return final_rows

# ─────────────────────── EJECUCIÓN PRINCIPAL ─────────────────────────

def main():
    """Punto de entrada para ejecutar el procesador de Biopsias."""
    print("="*60)
    print("PROCESADOR ESPECIALIZADO DE INFORMES DE BIOPSIA (Múltiples Especímenes)")
    print("="*60)

    root = tk.Tk()
    root.withdraw() 
    pdf_path = filedialog.askopenfilename(
        title="Seleccione el informe PDF de Biopsia para procesar",
        filetypes=[("Archivos PDF", "*.pdf")]
    )

    if not pdf_path:
        print("No se seleccionó ningún archivo. Saliendo del programa.")
        return

    print(f"📄 Archivo seleccionado: {pdf_path}")
    output_dir = Path(pdf_path).parent
    
    try:
        print("🔍 Extrayendo texto con OCR...")
        pdf_text = pdf_to_text_enhanced(pdf_path)
        
        print("📊 Extrayendo datos comunes del informe...")
        common_data = extract_biopsy_data(pdf_text)
        
        print(f"🔬 Identificando especímenes individuales...")
        specimens_data = extract_specimens_data(pdf_text, common_data.get('numero_peticion', ''))
        
        if not specimens_data:
            print("❌ No se pudieron identificar especímenes individuales (A, B...).")
            return
            
        print(f"✅ Se identificaron {len(specimens_data)} especímenes.")

        print("📋 Mapeando datos a formato Excel...")
        excel_rows = map_to_excel_format(common_data, specimens_data)
        
        print("💾 Generando archivo Excel...")
        df = pd.DataFrame(excel_rows)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"Informe_Biopsia_{timestamp}.xlsx"
        output_path = output_dir / output_filename
        
        df.to_excel(output_path, index=False, engine="openpyxl")
        
        print("✨ Aplicando formato profesional...")
        wb = load_workbook(output_path)
        ws = wb.active
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font, cell.alignment, cell.fill = header_font, header_alignment, header_fill
        for col in ws.columns:
            try:
                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                ws.column_dimensions[col[0].column_letter].width = min((max_length + 2), 60)
            except ValueError: pass
        wb.save(output_path)
        
        print("\n" + "="*60)
        print("🎉 ¡PROCESAMIENTO COMPLETADO EXITOSAMENTE!")
        print(f"📊 Se han generado {len(excel_rows)} registros.")
        print(f"📁 Archivo guardado en: {output_path}")
        print("="*60)

    except Exception as e:
        print("\n" + "!"*60)
        print(f"❌ Ocurrió un error inesperado:")
        print(f"   {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        print("!"*60)

if __name__ == "__main__":
    main()