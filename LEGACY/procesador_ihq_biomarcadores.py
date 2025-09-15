#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor dedicado de biomarcadores IHQ (HER2, Ki-67, ER/PR, PD-L1, P16, Estudios Solicitados).
No modifica el esquema operativo estándar ni otros procesadores.
Genera un Excel extendido (55 + 8 columnas) solo para los PDFs IHQ seleccionados.
Capaz de procesar PDFs con múltiples informes, generando una fila por cada uno.
"""

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ocr_processing import pdf_to_text_enhanced
import procesador_ihq as ihq


def _extract_biomarkers(text: str) -> dict:
    """
    Función de extracción de biomarcadores mejorada y robustecida para manejar
    múltiples variaciones de texto encontradas en los informes del HUV.
    """
    out = {
        "IHQ_HER2": "",
        "IHQ_KI-67": "",
        "IHQ_RECEPTOR_ESTROGENO": "",
        "IHQ_RECEPTOR_PROGESTAGENOS": "",
        "IHQ_PDL-1": "",
        "IHQ_ESTUDIOS_SOLICITADOS": "",
        "IHQ_P16_ESTADO": "",
        "IHQ_P16_PORCENTAJE": "",
    }

    # Estudios Solicitados (Patrón robustecido)
    m_est = re.search(r'(?i)estudios\s+solicitados:?\s*(?:se\s+realiz[oó]\s+tinci[oó]n\s+con\s+los\s+siguientes\s+marcadores|se\s+realiz[oó]\s+tinci[oó]n\s+especial\s+para\s*)?([A-Z0-9\s,+/.-]+)', text)
    if m_est:
        raw = m_est.group(1).strip()
        # Limpieza avanzada para evitar capturar texto no deseado
        raw = re.split(r'\n', raw)[0]
        tokens = []
        # Normaliza separadores comunes como coma, punto y coma, slash y espacios
        for token in re.split(r'[\s,;/]+', raw):
            t = token.strip().upper().replace(' ', '')
            if t and len(t) > 1: # Filtra tokens vacíos o de un solo caracter
                tokens.append(t)
        # Elimina duplicados manteniendo el orden
        out["IHQ_ESTUDIOS_SOLICITADOS"] = ', '.join(dict.fromkeys(tokens))

    # P16 (Patrones más flexibles)
    m_p16_estado = re.search(r'(?i)\bP\s*16\b[^\n]*?\b(positiv[oa]|negativ[oa])\b', text)
    if m_p16_estado:
        out["IHQ_P16_ESTADO"] = m_p16_estado.group(1).upper()
    elif re.search(r'(?i)\bP\s*16\b[^\n]*?\b(en\s+bloque|difus[ao])\b', text):
        out["IHQ_P16_ESTADO"] = 'POSITIVO'
    
    m_p16_pct = re.search(r'(?i)\bP\s*16\b[^\n%]*?(\d{1,3})\s*%', text)
    if m_p16_pct:
        out["IHQ_P16_PORCENTAJE"] = m_p16_pct.group(1)

    # HER2 (Lógica mejorada para score e ISH)
    m_her2 = re.search(r'(?i)\bHER2(?:/NEU)?\b\s*[:\-(]?\s*(0|1\+|2\+|3\+|negativ[oa]|positiv[oa])', text)
    her2_score = ''
    if m_her2:
        val = m_her2.group(1).upper()
        if 'NEGATIVO' in val: her2_score = 'NEGATIVO'
        elif 'POSITIVO' in val: her2_score = 'POSITIVO (3+)'
        else: her2_score = val

    m_ish = re.search(r'(?i)\b(?:ISH|FISH)\b[^\n]*?\b(amplificad[oa]|no\s*amplificad[oa])\b', text)
    her2_ish = m_ish.group(1).upper().replace(' ', '_') if m_ish else ''
    
    her2_final = her2_score
    if her2_ish:
        ish_status = 'AMPLIFICADO' if 'AMPLIFICAD' in her2_ish else 'NO AMPLIFICADO'
        her2_final = f"{her2_score} ({ish_status})" if her2_score else ish_status
    out["IHQ_HER2"] = her2_final.strip()

    # Ki-67 (Maneja errores de OCR y formatos de porcentaje)
    m_ki = re.search(r'(?i)\bK[IL]\s*[- ]?67\b[^\n%]*?(?:de|del|en|aproximadamente|menor del)?\s*<?\s*(\d{1,3})\s*%', text)
    if m_ki:
        out["IHQ_KI-67"] = f"{m_ki.group(1)}%"

    # ER/RE (Receptor de Estrógeno) - Patrón unificado y robusto
    m_re = re.search(r'(?i)(?:receptor(?:es)?(?:\s+hormonal)?\s+de\s+estr[oó]geno|\bRE\b|\bER\b)[^\n]*?(positiv[oa]|negativ[oa])?\s*(?:en\s+el|de|del)?\s*\(?(\d{1,3})\s*%\)?', text)
    re_estado, re_pct = '', ''
    if m_re:
        if m_re.group(1): re_estado = m_re.group(1).upper()
        if m_re.group(2): re_pct = m_re.group(2) + "%"
    else:
        m_re_estado_fallback = re.search(r'(?i)(?:receptor(?:es)?(?:\s+hormonal)?\s+de\s+estr[oó]geno|\bRE\b|\bER\b)[^\n]*?\b(positiv[oa]|negativ[oa])\b', text)
        m_re_pct_fallback = re.search(r'(?i)(?:receptor(?:es)?(?:\s+hormonal)?\s+de\s+estr[oó]geno|\bRE\b|\bER\b)[^\n%]*?(\d{1,3})\s*%', text)
        if m_re_estado_fallback: re_estado = m_re_estado_fallback.group(1).upper()
        if m_re_pct_fallback: re_pct = m_re_pct_fallback.group(1) + "%"
    out["IHQ_RECEPTOR_ESTROGENO"] = f"{re_estado} {re_pct}".strip()

    # PR/RP (Receptor de Progestágenos) - Patrón unificado y robusto
    m_rp = re.search(r'(?i)(?:receptor(?:es)?(?:\s+hormonal)?\s+de\s+progest(?:erona|ágenos)|\bRP\b|\bPR\b)[^\n]*?(positiv[oa]|negativ[oa])?\s*(?:en\s+el|de|del)?\s*\(?(\d{1,3})\s*%\)?', text)
    rp_estado, rp_pct = '', ''
    if m_rp:
        if m_rp.group(1): rp_estado = m_rp.group(1).upper()
        if m_rp.group(2): rp_pct = m_rp.group(2) + "%"
    else:
        m_rp_estado_fallback = re.search(r'(?i)(?:receptor(?:es)?(?:\s+hormonal)?\s+de\s+progest(?:erona|ágenos)|\bRP\b|\bPR\b)[^\n]*?\b(positiv[oa]|negativ[oa])\b', text)
        m_rp_pct_fallback = re.search(r'(?i)(?:receptor(?:es)?(?:\s+hormonal)?\s+de\s+progest(?:erona|ágenos)|\bRP\b|\bPR\b)[^\n%]*?(\d{1,3})\s*%', text)
        if m_rp_estado_fallback: rp_estado = m_rp_estado_fallback.group(1).upper()
        if m_rp_pct_fallback: rp_pct = m_rp_pct_fallback.group(1) + "%"
    out["IHQ_RECEPTOR_PROGESTAGENOS"] = f"{rp_estado} {rp_pct}".strip()
    
    # PD-L1 (Captura TPS y/o CPS)
    m_tps = re.search(r'(?i)\bPD\s*-?L1\b[^\n]*?\bTPS\b[^\n%<]*?<?\s*(\d{1,3})\s*%', text)
    m_cps = re.search(r'(?i)\bPD\s*-?L1\b[^\n]*?\bCPS\b[^\n\d<]*?<?\s*(\d{1,3})', text)
    pdl1_parts = []
    if m_tps: pdl1_parts.append(f"TPS {m_tps.group(1)}%")
    if m_cps: pdl1_parts.append(f"CPS {m_cps.group(1)}")
    out["IHQ_PDL-1"] = ' '.join(pdl1_parts)
    if not out["IHQ_PDL-1"]:
        m_pdl1_fallback = re.search(r'(?i)\bPD\s*-?L1\b\s*[:\-(]?\s*([^\n]+)', text)
        if m_pdl1_fallback:
            out["IHQ_PDL-1"] = m_pdl1_fallback.group(1).strip()

    return out


def process_ihq_paths(pdf_paths: list[str], output_dir: str) -> Path:
    rows = []
    extended_columns = None
    
    # ▼▼▼ INICIO DE LA MODIFICACIÓN PARA MÚLTIPLES INFORMES ▼▼▼
    
    for pdf in pdf_paths:
        full_text = pdf_to_text_enhanced(pdf)
        
        # Patrón para dividir el texto por cada informe. Busca el número de petición de IHQ.
        # El uso de 're.split' con un grupo de captura '(...)' mantiene el delimitador al inicio de cada fragmento.
        report_chunks = re.split(r'(N\.\s*peticion\s*:\s*IHQ\d+)', full_text, flags=re.IGNORECASE)

        # El primer elemento de 'report_chunks' es el texto antes de la primera coincidencia (si lo hay),
        # lo ignoramos. Luego, procesamos los fragmentos de dos en dos (delimitador + contenido).
        for i in range(1, len(report_chunks), 2):
            # Reconstruimos el texto del informe individual
            single_report_text = report_chunks[i] + report_chunks[i+1]
            
            # Aplicamos la misma lógica de extracción a este fragmento
            base = ihq.extract_ihq_data(single_report_text)
            base_rows = ihq.map_to_excel_format(base)
            
            if not base_rows:
                continue # Si no se extrae nada, pasamos al siguiente fragmento
                
            row = dict(base_rows[0])
            biomarkers = _extract_biomarkers(single_report_text)
            row.update(biomarkers)
            rows.append(row)

            # Se establece la estructura de columnas con el primer informe procesado exitosamente
            if extended_columns is None:
                base_cols = list(base_rows[0].keys())
                extra_cols = list(biomarkers.keys())
                extended_columns = base_cols + extra_cols
                
    # ▲▲▲ FIN DE LA MODIFICACIÓN ▲▲▲

    if not rows:
        raise RuntimeError("No se pudo extraer información de los PDFs IHQ seleccionados.")

    df = pd.DataFrame(rows)
    if extended_columns:
        df = df.reindex(columns=extended_columns)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = Path(output_dir) / f"Informe_IHQ_BIOMARCADORES_{timestamp}.xlsx"
    df.to_excel(output_path, index=False, engine='openpyxl')

    # Formato de encabezados
    wb = load_workbook(output_path)
    ws = wb.active
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
    for col in ws.columns:
        try:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = min((max_length + 2), 60)
        except ValueError:
            pass
    wb.save(output_path)
    return output_path