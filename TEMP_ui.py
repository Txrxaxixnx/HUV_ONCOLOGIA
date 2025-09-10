#ui.py:
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Interfaz grÃ¡fica del sistema OCR HUV."""

import threading
from datetime import datetime
from pathlib import Path
import configparser

import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ocr_processing import pdf_to_text_enhanced
from data_extraction import process_text_to_excel_rows, detect_report_type

_config = configparser.ConfigParser(interpolation=None)
_config.read(Path(__file__).resolve().parent / "config.ini", encoding="utf-8")

TIMESTAMP_FORMAT = _config.get("OUTPUT", "TIMESTAMP_FORMAT", fallback="%Y%m%d_%H%M%S")
OUTPUT_FILENAME = _config.get("OUTPUT", "OUTPUT_FILENAME", fallback="informes_medicos")

WINDOW_WIDTH = _config.getint("INTERFACE", "WINDOW_WIDTH", fallback=900)
WINDOW_HEIGHT = _config.getint("INTERFACE", "WINDOW_HEIGHT", fallback=700)
LOG_HEIGHT = _config.getint("INTERFACE", "LOG_HEIGHT", fallback=8)


class HUVOCRSystem:
    """Interfaz basada en Tkinter para procesar informes PDF."""

    def __init__(self, root):
        self.root = root
        root.title("Sistema OCR - Hospital Universitario del Valle")
        root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}")
        root.configure(bg="#f8f9fa")

        self.files = []
        self.output_dir = ""

        self._setup_gui()

    # ConfiguraciÃ³n de la interfaz
    def _setup_gui(self):
        title_frame = tk.Frame(self.root, bg="#f8f9fa")
        title_frame.pack(pady=15, fill="x")

        title_label = tk.Label(
            title_frame,
            text="SISTEMA OCR - HOSPITAL UNIVERSITARIO DEL VALLE",
            font=("Arial", 16, "bold"),
            bg="#f8f9fa",
            fg="#2c3e50",
        )
        title_label.pack()

        subtitle_label = tk.Label(
            title_frame,
            text="ExtracciÃ³n automatizada de informes de patologÃ­a",
            font=("Arial", 10),
            bg="#f8f9fa",
            fg="#7f8c8d",
        )
        subtitle_label.pack()

        button_frame = tk.Frame(self.root, bg="#f8f9fa")
        button_frame.pack(pady=10, fill="x", padx=20)

        tk.Button(
            button_frame,
            text="ðŸ“„ AÃ±adir PDFs",
            command=self.add_files,
            bg="#3498db",
            fg="white",
            padx=20,
            pady=10,
            font=("Arial", 10, "bold"),
        ).pack(side="left", padx=5)

        tk.Button(
            button_frame,
            text="ðŸ“ Carpeta PDFs",
            command=self.add_folder,
            bg="#2ecc71",
            fg="white",
            padx=20,
            pady=10,
            font=("Arial", 10, "bold"),
        ).pack(side="left", padx=5)

        tk.Button(
            button_frame,
            text="ðŸ—‘ï¸ Limpiar",
            command=self.clear_files,
            bg="#e74c3c",
            fg="white",
            padx=20,
            pady=10,
            font=("Arial", 10, "bold"),
        ).pack(side="right", padx=5)

        list_frame = tk.Frame(self.root, bg="#f8f9fa")
        list_frame.pack(fill="both", expand=True, padx=20, pady=5)

        list_label = tk.Label(
            list_frame,
            text="Archivos seleccionados:",
            font=("Arial", 12, "bold"),
            bg="#f8f9fa",
        )
        list_label.pack(anchor="w", pady=(0, 5))

        list_container = tk.Frame(list_frame)
        list_container.pack(fill="both", expand=True)

        self.file_listbox = tk.Listbox(
            list_container,
            font=("Consolas", 10),
            bg="white",
            selectmode=tk.EXTENDED,
            height=12,
        )
        scrollbar = tk.Scrollbar(list_container, orient="vertical")
        self.file_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.file_listbox.yview)
        self.file_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        output_frame = tk.Frame(self.root, bg="#f8f9fa")
        output_frame.pack(pady=15, fill="x", padx=20)

        tk.Button(
            output_frame,
            text="ðŸ“‚ Carpeta de Salida",
            command=self.select_output_dir,
            bg="#f39c12",
            fg="white",
            padx=15,
            pady=8,
            font=("Arial", 10, "bold"),
        ).pack(side="left", padx=5)

        self.output_label = tk.Label(
            output_frame,
            text="No seleccionada",
            bg="#f8f9fa",
            font=("Arial", 10),
        )
        self.output_label.pack(side="left", padx=15)

        process_frame = tk.Frame(self.root, bg="#f8f9fa")
        process_frame.pack(pady=15, fill="x", padx=20)

        self.progress_bar = ttk.Progressbar(process_frame, length=500, mode="determinate")
        self.progress_bar.pack(side="left", padx=10, expand=True, fill="x")

        tk.Button(
            process_frame,
            text="Procesar IHQ (Biomarcadores)",
            command=self.start_processing_ihq,
            bg="#16a085",
            fg="white",
            padx=15,
            pady=10,
            font=("Arial", 10, "bold"),
        ).pack(side="left", padx=5)

        tk.Button(
            process_frame,
            text="ðŸš€ PROCESAR INFORMES",
            command=self.start_processing,
            bg="#9b59b6",
            fg="white",
            padx=25,
            pady=12,
            font=("Arial", 12, "bold"),
        ).pack(side="right", padx=10)

        log_frame = tk.Frame(self.root, bg="#f8f9fa")
        log_frame.pack(fill="both", expand=True, padx=20, pady=(5, 20))

        log_label = tk.Label(
            log_frame,
            text="Registro de procesamiento:",
            font=("Arial", 12, "bold"),
            bg="#f8f9fa",
        )
        log_label.pack(anchor="w", pady=(0, 5))

        log_container = tk.Frame(log_frame)
        log_container.pack(fill="both", expand=True)

        self.log_text = tk.Text(
            log_container,
            height=LOG_HEIGHT,
            bg="#2c3e50",
            fg="#ecf0f1",
            font=("Consolas", 9),
            wrap="word",
        )
        log_scrollbar = tk.Scrollbar(log_container, orient="vertical")
        self.log_text.config(yscrollcommand=log_scrollbar.set)
        log_scrollbar.config(command=self.log_text.yview)
        self.log_text.pack(side="left", fill="both", expand=True)
        log_scrollbar.pack(side="right", fill="y")

    # Manejo de archivos
    def add_files(self):
        files = filedialog.askopenfilenames(
            title="Seleccionar informes PDF",
            filetypes=[("Archivos PDF", "*.pdf"), ("Todos los archivos", "*.*")],
        )
        added = 0
        for file_path in files:
            if file_path not in self.files:
                self.files.append(file_path)
                filename = Path(file_path).name
                self.file_listbox.insert(tk.END, f"ðŸ“„ {filename}")
                added += 1
        self._log(f"âž• {added} archivos aÃ±adidos. Total: {len(self.files)}")

    def add_folder(self):
        folder = filedialog.askdirectory(title="Seleccionar carpeta con PDFs")
        if not folder:
            return
        pdf_files = list(Path(folder).glob("*.pdf"))
        added = 0
        for pdf_path in pdf_files:
            file_str = str(pdf_path)
            if file_str not in self.files:
                self.files.append(file_str)
                self.file_listbox.insert(tk.END, f"ðŸ“„ {pdf_path.name}")
                added += 1
        self._log(f"ðŸ“ {added} archivos aÃ±adidos desde carpeta. Total: {len(self.files)}")

    def clear_files(self):
        self.files.clear()
        self.file_listbox.delete(0, tk.END)
        self._log("ðŸ—‘ï¸ Lista de archivos limpiada")

    def select_output_dir(self):
        directory = filedialog.askdirectory(title="Seleccionar carpeta de salida")
        if directory:
            self.output_dir = directory
            self.output_label.config(text=f"ðŸ“ {directory}")
            self._log(f"ðŸ“‚ Carpeta de salida: {directory}")

    # Logs y procesamiento
    def _log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def start_processing(self):
        if not self.files:
            messagebox.showwarning("Sin archivos", "Seleccione al menos un archivo PDF para procesar.")
            return
        if not self.output_dir:
            messagebox.showwarning("Sin carpeta de salida", "Seleccione una carpeta para guardar los resultados.")
            return
        threading.Thread(target=self._process_files, daemon=True).start()

    def _process_files(self):
        total_files = len(self.files)
        self.progress_bar["maximum"] = total_files
        all_rows = []
        processed_count = 0
        error_count = 0
        self._log(f"ðŸš€ Iniciando procesamiento de {total_files} archivos")
        self._log("=" * 60)
        for idx, pdf_path in enumerate(self.files, 1):
            filename = Path(pdf_path).name
            self._log(f"ðŸ“„ [{idx}/{total_files}] Procesando: {filename}")
            try:
                self._log("   ðŸ” Extrayendo texto con OCR...")
                pdf_text = pdf_to_text_enhanced(pdf_path)

                # ----- INICIO DE CÃ“DIGO DE DEPURACIÃ“N -----
                # Guardar el texto crudo del OCR en un archivo para anÃ¡lisis
                debug_filename = f"DEBUG_OCR_OUTPUT_{filename}.txt"
                try:
                    with open(Path(self.output_dir) / debug_filename, "w", encoding="utf-8") as f:
                        f.write(pdf_text)
                    self._log(f"   ðŸ› Â¡Texto de OCR guardado en {debug_filename} para depuraciÃ³n!")
                except Exception as e:
                    self._log(f"   âŒ No se pudo guardar el archivo de depuraciÃ³n: {e}")
                # ----- FIN DE CÃ“DIGO DE DEPURACIÃ“N -----

                if not pdf_text.strip():
                    self._log("   âš ï¸  Advertencia: No se extrajo texto del PDF")
                    continue
                self._log("   ðŸ“Š Extrayendo datos estructurados...")
                tipo_informe = detect_report_type(pdf_text)
                excel_rows = process_text_to_excel_rows(pdf_text, filename)
                extracted_data = {'tipo_informe': tipo_informe, 'specimens': [None] * len(excel_rows)}
                # ----- INICIO DE CÃ“DIGO DE DEPURACIÃ“N DE MALIGNIDAD -----
                print(f"\n--- DEBUG: {filename} ---")
                diagnostico_texto = extracted_data.get('diagnostico', 'Â¡Â¡Â¡DIAGNÃ“STICO NO ENCONTRADO!!!')
                microscopica_texto = extracted_data.get('descripcion_microscopica', 'Â¡Â¡Â¡DESCRIPCIÃ“N MICROSCÃ“PICA NO ENCONTRADA!!!')
                
                print(f"TEXTO DEL DIAGNÃ“STICO EXTRAÃDO:\n---\n{diagnostico_texto}\n---")
                print(f"TEXTO MICROSCÃ“PICO EXTRAÃDO:\n---\n{microscopica_texto}\n---")
                print(f"RESULTADO DE MALIGNIDAD CALCULADO: {extracted_data.get('malignidad')}")
                print("--- FIN DEBUG ---\n")
                # ----- FIN DE CÃ“DIGO DE DEPURACIÃ“N -----
                fuente = extracted_data.get('fecha_ordenamiento_fuente', '')
                if fuente:
                    self._log(f"   ðŸ—“ï¸ Fecha ordenamiento desde: {fuente}")
                if extracted_data.get('eps_normalizado'):
                    self._log("   ðŸ”§ EPS normalizada")
                if extracted_data.get('servicio_normalizado'):
                    self._log("   ðŸ”§ Servicio normalizado")
                self._log("   ðŸ“‹ Mapeando a formato Excel...")
                map_to_excel_format = lambda *args, **kwargs: excel_rows
                excel_rows = map_to_excel_format(extracted_data, filename)
                all_rows.extend(excel_rows)
                processed_count += 1
                tipo_informe = extracted_data.get('tipo_informe', 'DESCONOCIDO')
                num_specimens = len(extracted_data.get('specimens', []))
                self._log(f"   âœ… Completado - Tipo: {tipo_informe} - EspecÃ­menes: {num_specimens}")
            except Exception as e:
                error_count += 1
                self._log(f"   âŒ Error: {str(e)}")
            self.progress_bar["value"] = idx
            self.root.update_idletasks()
        if all_rows:
            self._log("=" * 60)
            self._log("ðŸ’¾ Generando archivo Excel...")
            try:
                df = pd.DataFrame(all_rows)
                timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)
                output_filename = f"{OUTPUT_FILENAME}_{timestamp}.xlsx"
                output_path = Path(self.output_dir) / output_filename
                df.to_excel(output_path, index=False, engine="openpyxl")
                # --- INICIO DE CÃ“DIGO PARA FORMATEAR ENCABEZADOS ---
                self._log(" Â  âœ¨ Aplicando formato profesional a encabezados...")
                wb = load_workbook(output_path)
                ws = wb.active

                # Definir estilos: fuente profesional y ajuste de texto
                header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

                # Aplicar estilos a cada celda del encabezado (primera fila)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill

                # Ajustar el ancho de las columnas para una mejor visualizaciÃ³n
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = min(adjusted_width, 50) # Limitar ancho mÃ¡ximo

                # Guardar el archivo Excel con el nuevo formato
                wb.save(output_path)
                
                self._log("=" * 60)
                self._log("ðŸŽ‰ PROCESAMIENTO COMPLETADO")
                self._log(f"âœ… Archivos procesados exitosamente: {processed_count}")
                self._log(f"âŒ Archivos con errores: {error_count}")
                self._log(f"ðŸ“Š Total de registros generados: {len(all_rows)}")
                self._log(f"ðŸ“ Archivo guardado: {output_filename}")
                self._log("=" * 60)
                mensaje = "\n".join([
                    "âœ… Procesamiento exitoso!",
                    "",
                    f"ðŸ“Š {processed_count} archivos procesados",
                    f"ðŸ“„ {len(all_rows)} registros generados",
                    f"ðŸ“ Archivo: {output_filename}",
                    "",
                    f"El archivo Excel ha sido guardado en:\n{output_path}",
                ])
                messagebox.showinfo("Procesamiento Completado", mensaje)
            except Exception as e:
                self._log(f"âŒ Error generando Excel: {str(e)}")
                messagebox.showerror("Error", f"Error generando archivo Excel:\n{str(e)}")
        else:
            self._log("âš ï¸ No se procesÃ³ ningÃºn archivo exitosamente")
            messagebox.showwarning("Sin resultados", "No se pudo extraer informaciÃ³n de ningÃºn archivo")

