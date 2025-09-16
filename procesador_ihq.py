# procesador_ihq.py:
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesador independiente y especializado para informes de Inmunohistoqu√≠mica (IHQ) del HUV.

Este script contiene toda la l√≥gica y patrones necesarios para extraer datos
de los PDFs de IHQ, garantizando precisi√≥n al no depender de constantes gen√©ricas.
"""

import os
import re
import sys
import unicodedata
from pathlib import Path
from datetime import datetime, date
from dateutil.relativedelta import relativedelta

# --- Dependencias requeridas (pip install PyMuPDF pillow pytesseract pandas openpyxl python-dateutil) ---
import pandas as pd
import fitz  # PyMuPDF
from PIL import Image
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tkinter as tk
from tkinter import filedialog

# --- M√≥dulos del proyecto necesarios ---
# Se asume que este script puede encontrar los siguientes archivos en su entorno.
try:
    from ocr_processing import pdf_to_text_enhanced
    # Importamos solo las configuraciones y constantes GEN√âRICAS que s√≠ aplican
    from huv_constants import HUV_CONFIG, CUPS_CODES, PROCEDIMIENTOS, PATTERNS_HUV
except ImportError:
    print("ERROR: Aseg√∫rese de que 'ocr_processing.py' y 'huv_constants.py' est√©n accesibles.")
    sys.exit(1)

# ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ CONSTANTES ESPECIALIZADAS PARA IHQ ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

# Palabras clave de malignidad extendidas espec√≠ficamente para informes de IHQ
MALIGNIDAD_KEYWORDS_IHQ = [
    'CARCINOMA', 'CANCER', 'MALIGNO', 'MALIGNIDAD', 'METASTASIS', 'METAST√ÅSICO',
    'NEOPLASIA MALIGNA', 'TUMOR MALIGNO', 'ADENOCARCINOMA', 'LINFOMA',
    'SARCOMA', 'MELANOMA', 'LEUCEMIA', 'HODGKIN', 'HODKING', 'PAGET'
]

# Patrones Regex dise√±ados y ajustados para AMBAS plantillas de IHQ
PATTERNS_IHQ = {
    # CORREGIDO: tolerante a acentos y tambi√©n cuando el PDF trae solo ‚ÄúDESCRIPCION MACROSCOPICA‚Äù
    'descripcion_macroscopica_ihq': r'DESCRIPCI[√ìO]N\s+MACROSC[√ìO]PICA([\s\S]+?)(?=DESCRIPCI[√ìO]N\s+MICROSC[√ìO]PICA|RESULTADO\s+DE\s+INMUNOHISTOQU[√çI]MICA)',
    
    # CORREGIDO: idem para el bloque microsc√≥pico
    'descripcion_microscopica_ihq': r'(?:DESCRIPCI[√ìO]N\s+MICROSC[√ìO]PICA|RESULTADO\s+DE\s+INMUNOHISTOQU[√çI]MICA\.?)([\s\S]+?)(?=\n\s*DIAGN[√ìO]STICO\s*\n)',

    'diagnostico_final_ihq': r'DIAGN√ìSTICO\s*\n([\s\S]+?)(?=\s*(?:ARMANDO CORTES BUELVAS\s*\n\s*Responsable del an√°lisis|NANCY MEJIA VARGAS\s*\n\s*M√©dica Pat√≥loga)|Nota: Este informe)',
    
    # Nuevo patr√≥n para encontrar la fecha de diagn√≥stico en el segundo tipo de informe.
    'fecha_diagnostico_ihq': r'Fecha de diagn√≥stico\s*:\s*([\d\-/]+)',
    
    # Patr√≥n m√°s gen√©rico para el responsable, busca un nombre encima de su t√≠tulo.
    'responsable_ihq': r'([A-Z\s]+)\n\s*(?:M√©dica Pat√≥loga|Responsable del an√°lisis)',
}
# (En la secci√≥n de FUNCIONES DE UTILIDAD)

def split_full_name(full_name: str) -> dict:
    """CORREGIDO: Maneja correctamente nombres de 2, 3 y 4+ partes."""
    if not full_name: return {'primer_nombre': '', 'segundo_nombre': '', 'primer_apellido': '', 'segundo_apellido': ''}
    parts = [p.strip() for p in full_name.strip().split() if p.strip()]
    result = {'primer_nombre': '', 'segundo_nombre': '', 'primer_apellido': '', 'segundo_apellido': ''}
    if len(parts) == 1:
        result['primer_nombre'] = parts[0]
    elif len(parts) == 2: # LILIA ROJAS -> LILIA, ROJAS
        result['primer_nombre'], result['primer_apellido'] = parts[0], parts[1]
    elif len(parts) == 3: # LILIA MARIA ROJAS -> LILIA, MARIA, ROJAS
        result['primer_nombre'], result['segundo_nombre'], result['primer_apellido'] = parts[0], parts[1], parts[2]
    elif len(parts) >= 4: # JEISON ARMANDO RIVERA HENAO
        result['primer_nombre'], result['segundo_nombre'], result['primer_apellido'] = parts[0], parts[1], parts[2]
        result['segundo_apellido'] = ' '.join(parts[3:])
    return result

def calculate_birth_date(edad_str: str, fecha_referencia_str: str = None) -> str:
    """CORREGIDO: Implementaci√≥n precisa que usa a√±os, meses y d√≠as."""
    try:
        years_match = re.search(r'(\d+)\s*a[√±n]os', edad_str, re.IGNORECASE)
        months_match = re.search(r'(\d+)\s*mes(es)?', edad_str, re.IGNORECASE)
        days_match = re.search(r'(\d+)\s*d[i√≠]as', edad_str, re.IGNORECASE)
        
        years = int(years_match.group(1)) if years_match else 0
        months = int(months_match.group(1)) if months_match else 0
        days = int(days_match.group(1)) if days_match else 0

        ref_date = date.today()
        if fecha_referencia_str:
            for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
                try:
                    ref_date = datetime.strptime(fecha_referencia_str, fmt).date()
                    break
                except ValueError: continue
        
        birth_date = ref_date - relativedelta(years=years, months=months, days=days)
        return birth_date.strftime('%d/%m/%Y')
    except Exception:
        return ''

def convert_date_format(date_str: str) -> str:
    if not date_str: return ''
    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime('%d/%m/%Y')
        except ValueError: pass
    return date_str

def _normalize_text(u: str) -> str:
    return unicodedata.normalize('NFKD', u or '').encode('ASCII', 'ignore').decode().upper()

def detect_malignancy_ihq(*texts: str) -> str:
    """CORREGIDO: Usa la lista de palabras clave espec√≠fica de IHQ."""
    text_to_check = " ".join(_normalize_text(t) for t in texts if t)
    for keyword in MALIGNIDAD_KEYWORDS_IHQ:
        if re.search(r'\b' + keyword + r'\b', text_to_check):
            return 'PRESENTE'
    return 'AUSENTE'

def deduce_specialty_ihq(servicio: str) -> str:
    # L√≥gica de especialidad, puede ser tan espec√≠fica como se necesite
    if 'GINECOLOGIA' in servicio.upper():
        return 'GINECOLOGIA Y OBSTETRICIA'
    return 'MEDICINA GENERAL'

# ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ EXTRACCI√ìN Y MAPEO ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
# (En la secci√≥n de EXTRACCI√ìN Y MAPEO)

"""
Fase 2 ‚Äì Prototipos (IHQ) para nuevos biomarcadores y campos

Objetivo
- Servir como gu√≠a de implementaci√≥n para extraer valores de biomarcadores a partir del texto OCR, coherente con analisis/01_ESQUEMA_DE_DATOS.md.

Ubicaci√≥n sugerida de b√∫squeda
- Priorizar secciones ‚ÄúRESULTADO DE INMUNOHISTOQU√çMICA‚Äù y ‚ÄúDESCRIPCI√ìN MICROSC√ìPICA‚Äù.
- Usar fallback al documento completo si no se encuentra en secciones prioritarias.

Normalizaci√≥n general
- Ignorar acentos y may√∫sculas (usar `re.IGNORECASE` y normalizaci√≥n).
- Unificar alias: "SOX 10"‚Üí"SOX10", "PD L1"‚Üí"PD-L1", "Ki 67"‚Üí"Ki-67".

Regex/heur√≠sticas propuestas
- Estudios Solicitados:
  - Patr√≥n: `(?i)estudios\s+solicitados:?\s*(?:se\s+realiz[o√≥]\s+tinci[o√≥]n\s+especial\s+para\s*)?([A-Z0-9 ,+/\.-]+)`
  - Post: dividir por comas/espacios, limpiar tokens, devolver lista normalizada.
- P16 (Estado/Porcentaje):
  - Estado: `(?i)\bP\s*16\b[^\n]*\b(positiv[oa]|negativ[oa])\b`
  - Porcentaje: `(?i)\bP\s*16\b[^\n%]*?(\d{1,3})\s*%`
  - Heur√≠stica: si aparece ‚Äúen bloque/difusa‚Äù, tratar como POSITIVO.
- HER2 (score ¬± ISH/FISH):
  - Score: `(?i)\bHER2(?:/neu)?\b\s*[:\-]?\s*(0|1\+|2\+|3\+)`
  - ISH/FISH: `(?i)\b(?:ISH|FISH)\b[^\n]*\b(amplificad[oa]|no\s*amplificad[oa])\b`
- Ki-67 (porcentaje):
  - `(?i)\bK[il]\s*[- ]?67\b[^\n%]*?(\d{1,3})\s*%`
- RE/ER (estado y %):
  - Estado: `(?i)(receptor(?:\s+hormonal)?\s+de\s+estr(?:o|√≥)geno|\bER\b)[^\n]*\b(positiv[oa]|negativ[oa])\b`
  - %: `(?i)(receptor(?:\s+hormonal)?\s+de\s+estr(?:o|√≥)geno|\bER\b)[^\n%]*?(\d{1,3})\s*%`
- RP/PR (estado y %):
  - Estado: `(?i)(receptor(?:\s+hormonal)?\s+de\s+progest(?:erona|√°genos)|\bPR\b)[^\n]*\b(positiv[oa]|negativ[oa])\b`
  - %: `(?i)(receptor(?:\s+hormonal)?\s+de\s+progest(?:erona|√°genos)|\bPR\b)[^\n%]*?(\d{1,3})\s*%`
- PD-L1 (TPS/CPS):
  - TPS: `(?i)\bPD\s*-?L1\b[^\n]*\bTPS\b[^\n%]*?(\d{1,3})\s*%`
  - CPS: `(?i)\bPD\s*-?L1\b[^\n]*\bCPS\b[^\n\d]*?(\d{1,3})`

Sugerencia de claves en `extracted_data`
- `ihq_estudios_solicitados` ‚Üí list[str]
- `ihq_p16_estado` ‚Üí {POSITIVO|NEGATIVO|""}
- `ihq_p16_porcentaje` ‚Üí "0-100" (texto)
- `ihq_her2_score` ‚Üí {0,1+,2+,3+}
- `ihq_her2_ish` ‚Üí {AMPLIFICADO|NO AMPLIFICADO|""}
- `ihq_ki67_porcentaje` ‚Üí "0-100"
- `ihq_re_estado`, `ihq_re_porcentaje`
- `ihq_rp_estado`, `ihq_rp_porcentaje`
- `ihq_pdl1_tps_porcentaje`, `ihq_pdl1_cps`

Validaci√≥n
- Verificar coherencia (p. ej., 0‚â§%‚â§100). Para P16, si hay estado pero sin %, dejar porcentaje vac√≠o.
- Registrar vac√≠o cuando no se mencione expl√≠citamente en el texto.
"""

def extract_ihq_data(text: str) -> dict:
    """Funci√≥n orquestadora mejorada para la extracci√≥n de datos de AMBOS tipos de IHQ."""
    data = {}
    
    # 1. Extracci√≥n b√°sica con patrones comunes
    for key, pattern in PATTERNS_HUV.items():
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        data[key] = re.sub(r'\s+', ' ', match.group(1).strip()) if match else ''

    # 2. Extracci√≥n de bloques con los patrones flexibles de IHQ
    for key, pattern in PATTERNS_IHQ.items():
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        data[key] = match.group(1).strip() if match else ''

    # 3. Normalizaci√≥n y L√≥gica de Negocio Adaptativa
    data['tipo_informe'] = 'INMUNOHISTOQUIMICA'
    if data.get('eps', '').strip() == '5.0.5':
        data['eps'] = 'S.O.S'
        
    if data.get('nombre_completo'):
        data.update(split_full_name(data['nombre_completo']))
        
    if data.get('identificacion_numero'):
        data['identificacion_numero'] = re.sub(r'[^\d]', '', data['identificacion_numero'])
        
    if data.get('edad'):
        data['fecha_nacimiento'] = calculate_birth_date(data['edad'], convert_date_format(data.get('fecha_ingreso')))
        data['edad'] = re.search(r'(\d+)', data['edad']).group(1) if re.search(r'(\d+)', data['edad']) else ''
    
    # L√≥gica para encontrar al responsable (funciona para ambos casos)
    resp_name_match = re.search(PATTERNS_IHQ['responsable_ihq'], text, re.IGNORECASE)
    resp_name = resp_name_match.group(1).strip() if resp_name_match else data.get('responsable_analisis', '')
    
    if 'NANCY MEJIA' in resp_name.upper():
        data['responsable_final'] = 'NANCY MEJIA'
    elif 'ARMANDO CORTES' in resp_name.upper():
        data['responsable_final'] = 'ARMANDO CORTES'
    else:
        data['responsable_final'] = resp_name

    data['hospitalizado'] = 'NO'
    data['n_autorizacion'] = 'COEX'
    data['identificador_unico'] = '0'
    data['especialidad_deducida'] = deduce_specialty_ihq(data.get('servicio', ''))
    data['malignidad'] = detect_malignancy_ihq(data.get('diagnostico_final_ihq', ''))
    
    # L√≥gica para prefijos de descripci√≥n
    if "Se recibe orden" in text:
        prefix_macro = "Se recibe orden para realizaci√≥n de inmunohistoqu√≠mica"
        data['descripcion_macroscopica_final'] = (prefix_macro + data.get('descripcion_macroscopica_ihq', '')).strip()
    else:
        prefix_macro = "Informe de Estudios de Inmunohistoqu√≠mica"
        data['descripcion_macroscopica_final'] = (prefix_macro + data.get('descripcion_macroscopica_ihq', '')).strip()

    data['descripcion_microscopica_final'] = data.get('descripcion_microscopica_ihq', '').strip()
    data['diagnostico_final'] = data.get('diagnostico_final_ihq', '').strip()
    # Elimina espec√≠ficamente el texto basura del OCR y cualquier l√≠nea vac√≠a resultante.
    data['diagnostico_final'] = re.sub(r'\s*\.?\s*Nanty T‚ÄúM a U', '', data['diagnostico_final']).strip()

    # L√≥gica para encontrar el √≥rgano (tolerante a OCR y variantes de redacci√≥n)
    organo_final = ''

    macro_txt = data.get('descripcion_macroscopica_final', '') or ''
    macro_norm = _normalize_text(macro_txt)

    # 1) ‚ÄúR√ìTULO CORRESPONDIENTE A "..."‚Äù o ‚ÄúCORRESPONDE/ CORRESPONDIENTE A "..."‚Äù
    for pat in [
        r'ROTULO\s+CORRESPONDIENTE\s+A\s*["‚Äú]?([^"‚Äù\n]+)',
        r'CORRESPONDIENTE\s+A\s*["‚Äú]?([^"‚Äù\n]+)',
        r'CORRESPONDE\s+A\s*["‚Äú]?([^"‚Äù\n]+)',
        r'["‚Äú]([^"‚Äù\n]+)["‚Äù]\s+CON\s+DIAGN[O√ì]STICO'  # ‚Äú... ‚Äù con diagn√≥stico ...
    ]:
        m = re.search(pat, macro_norm, re.IGNORECASE)
        if m:
            organo_final = m.group(1).strip(' ."‚Äù')
            break

    # 2) Fallback: tabla ‚ÄúEstudios solicitados ‚Ä¶  Organo‚Äù
    if not organo_final:
        # Captura la l√≠nea inmediatamente posterior al encabezado con la palabra ORGANO
        mrow = re.search(r'(?i)ALMACENAMIENTO[^\n]*ORGANO[^\n]*\n([^\n]+)', text)
        if mrow:
            # Divide por grandes espacios entre columnas y toma la celda de √≥rgano
            parts = re.split(r'\s{2,}', mrow.group(1).strip())
            if len(parts) >= 2:
                organo_final = parts[1].strip()

    data['organo_final'] = organo_final if organo_final else 'No especificado'

    # L√≥gica para encontrar la fecha de ordenamiento
    fecha_diag_match = re.search(PATTERNS_IHQ['fecha_diagnostico_ihq'], text, re.IGNORECASE)
    if fecha_diag_match:
        data['fecha_ordenamiento'] = convert_date_format(fecha_diag_match.group(1))
    else:
        # Valor por defecto o buscar otra fecha si es necesario. Para el primer caso, se usaba un valor fijo.
        # Aqu√≠ puedes decidir qu√© hacer si no se encuentra. Por ahora, lo dejamos vac√≠o si no lo halla.
        data['fecha_ordenamiento'] = "28/07/2025" # Mantenemos el valor para el primer caso si no se halla otro.

    return data

# (Reemplazar esta funci√≥n completa)

def map_to_excel_format(extracted_data: dict) -> list:
    """Mapea los datos extra√≠dos al formato final de 55 columnas del HUV."""
    row_data = {}

    row_data["N. peticion (0. Numero de biopsia)"] = extracted_data.get('numero_peticion', '')
    row_data["Hospitalizado"] = extracted_data.get('hospitalizado', 'NO')
    row_data["Sede"] = HUV_CONFIG['sede_default']
    row_data["EPS"] = extracted_data.get('eps', '')
    row_data["Servicio"] = extracted_data.get('servicio', '')
    row_data["M√©dico tratante"] = extracted_data.get('medico_tratante', '')
    row_data["Especialidad"] = extracted_data.get('especialidad_deducida', '')
    # CORREGIDO: La ubicaci√≥n ahora se toma del servicio, es din√°mica.
    row_data["Ubicaci√≥n"] = extracted_data.get('servicio', '')
    row_data["N. Autorizacion"] = extracted_data.get('n_autorizacion', '')
    row_data["Identificador Unico"] = extracted_data.get('identificador_unico', '')
    # L√≥gica para Datos Cl√≠nicos
    if "TUMOR MALIGNO" in extracted_data.get('descripcion_macroscopica_final', ''):
        row_data["Datos Clinicos"] = 'NO'
    else:
        row_data["Datos Clinicos"] = 'SI'
    # CORREGIDO: La fecha de ordenamiento es ahora din√°mica.
    row_data["Fecha ordenamiento"] = extracted_data.get('fecha_ordenamiento', '')
    row_data["Tipo de documento"] = extracted_data.get('tipo_documento', HUV_CONFIG['tipo_documento_default'])
    row_data["N. de identificaci√≥n"] = extracted_data.get('identificacion_numero', '')
    row_data["Primer nombre"] = extracted_data.get('primer_nombre', '')
    row_data["Segundo nombre"] = extracted_data.get('segundo_nombre', '')
    row_data["Primer apellido"] = extracted_data.get('primer_apellido', '')
    row_data["Segundo apellido"] = extracted_data.get('segundo_apellido', '')
    row_data["Fecha de nacimiento"] = extracted_data.get('fecha_nacimiento', '')
    row_data["Edad"] = extracted_data.get('edad', '')
    row_data["Genero"] = extracted_data.get('genero', '')
    row_data["Departamento"] = HUV_CONFIG['departamento_default']
    row_data["Municipio"] = HUV_CONFIG['municipio_default']
    row_data["N. muestra"] = extracted_data.get('numero_peticion', '')
    row_data["CUPS"] = CUPS_CODES.get('INMUNOHISTOQUIMICA', '')
    row_data["Tipo de examen (4, 12, Metodo de obtenci√≥n de la muestra, factor de certeza para el diagn√≥stico)"] = "ESTUDIO DE INMUNOHISTOQUIMICA"
    row_data["Procedimiento (11. Tipo de estudio para el diagn√≥stico)"] = PROCEDIMIENTOS.get(CUPS_CODES.get('INMUNOHISTOQUIMICA'), '')
    row_data["Organo (1. Muestra enviada a patolog√≠a)"] = extracted_data.get('organo_final', '')
    row_data["Tarifa"] = HUV_CONFIG['tarifa_default']
    row_data["Valor"] = HUV_CONFIG['valor_default']
    row_data["Fecha de ingreso (2. Fecha de la muestra)"] = convert_date_format(extracted_data.get('fecha_ingreso', ''))
    row_data["Fecha finalizacion (3. Fecha del informe)"] = convert_date_format(extracted_data.get('fecha_informe', ''))
    
    resp_name = extracted_data.get('responsable_final', '')
    row_data["Usuario finalizacion"] = resp_name
    row_data["Malignidad"] = extracted_data.get('malignidad', 'AUSENTE')
    
    row_data["Descripcion macroscopica"] = extracted_data.get('descripcion_macroscopica_final', '')
    row_data["Descripcion microscopica (8,9, 10,12,. Invasi√≥n linfovascular y perineural, indice mit√≥tico/Ki67, Inmunohistoqu√≠mica, tama√±o tumoral)"] = extracted_data.get('descripcion_microscopica_final', '')
    row_data["Descripcion Diagnostico (5,6,7 Tipo histol√≥gico, subtipo histol√≥gico, margenes tumorales)"] = extracted_data.get('diagnostico_final', '')
    row_data["Diagnostico Principal"] = ""
    
    row_data["Hora Desc. macro"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    row_data["Responsable macro"] = resp_name
    
    EXCEL_COLUMNS = [
        "N. peticion (0. Numero de biopsia)", "Hospitalizado", "Sede", "EPS", "Servicio",
        "M√©dico tratante", "Especialidad", "Ubicaci√≥n", "N. Autorizacion", "Identificador Unico",
        "Datos Clinicos", "Fecha ordenamiento", "Tipo de documento", "N. de identificaci√≥n",
        "Primer nombre", "Segundo nombre", "Primer apellido", "Segundo apellido", "Fecha de nacimiento",
        "Edad", "Genero", "N√∫mero celular", "Direccion de correo electronico", "Direccion de correo electronico 2",
        "Contacto de emergencia", "Departamento", "Tel√©fono del contacto", "Municipio", "N. muestra",
        "CUPS", "Tipo de examen (4, 12, Metodo de obtenci√≥n de la muestra, factor de certeza para el diagn√≥stico)",
        "Procedimiento (11. Tipo de estudio para el diagn√≥stico)", "Organo (1. Muestra enviada a patolog√≠a)",
        "Tarifa", "Valor", "Copago", "Descuento", "Fecha de ingreso (2. Fecha de la muestra)",
        "Fecha finalizacion (3. Fecha del informe)", "Usuario finalizacion", "Usuario asignacion micro",
        "Fecha asignacion micro", "Malignidad", "Condicion", "Descripcion macroscopica",
        "Descripcion microscopica (8,9, 10,12,. Invasi√≥n linfovascular y perineural, indice mit√≥tico/Ki67, Inmunohistoqu√≠mica, tama√±o tumoral)",
        "Descripcion Diagnostico (5,6,7 Tipo histol√≥gico, subtipo histol√≥gico, margenes tumorales)",
        "Diagnostico Principal", "Comentario", "Informe adicional", "Congelaciones /Otros estudios",
        "Liquidos (5 Tipo histologico)", "Citometria de flujo (5 Tipo histologico)", "Hora Desc. macro",
        "Responsable macro"
    ]
    final_row = {col: row_data.get(col, '') for col in EXCEL_COLUMNS}

    return [final_row]

# ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ EJECUCI√ìN PRINCIPAL ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ

def main():
    """Punto de entrada para ejecutar el procesador de forma independiente."""
    print("="*60)
    print("PROCESADOR ESPECIALIZADO DE INFORMES DE IHQ - HUV")
    print("="*60)

    root = tk.Tk()
    root.withdraw() 
    pdf_path = filedialog.askopenfilename(
        title="Seleccione el informe PDF de IHQ para procesar",
        filetypes=[("Archivos PDF", "*.pdf")]
    )

    if not pdf_path:
        print("No se seleccion√≥ ning√∫n archivo. Saliendo del programa.")
        return

    print(f"üìÑ Archivo seleccionado: {pdf_path}")
    output_dir = Path(pdf_path).parent
    
    try:
        print("üîç Extrayendo texto con OCR... (esto puede tardar un momento)")
        pdf_text = pdf_to_text_enhanced(pdf_path)
        
        print("üìä Aplicando l√≥gica de extracci√≥n especializada para IHQ...")
        extracted_data = extract_ihq_data(pdf_text)
        
        print("üìã Mapeando datos a formato Excel...")
        excel_rows = map_to_excel_format(extracted_data)
        
        if not excel_rows:
            print("‚ùå No se pudo extraer informaci√≥n v√°lida del PDF.")
            return

        print("üíæ Generando archivo Excel...")
        df = pd.DataFrame(excel_rows)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"Informe_IHQ_{timestamp}.xlsx"
        output_path = output_dir / output_filename
        
        df.to_excel(output_path, index=False, engine="openpyxl")
        
        print("‚ú® Aplicando formato profesional a encabezados...")
        wb = load_workbook(output_path)
        ws = wb.active
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for cell in ws[1]:
            cell.font, cell.alignment, cell.fill = header_font, header_alignment, header_fill

        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col[0].column_letter].width = min(adjusted_width, 60)

        wb.save(output_path)
        
        print("\n" + "="*60)
        print("üéâ ¬°PROCESAMIENTO COMPLETADO EXITOSAMENTE!")
        print(f"‚úÖ Se ha generado 1 registro.")
        print(f"üìÅ Archivo guardado en: {output_path}")
        print("="*60)

    except Exception as e:
        print("\n" + "!"*60)
        print(f"‚ùå Ocurri√≥ un error inesperado durante el procesamiento:")
        print(f"   {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        print("!"*60)

if __name__ == "__main__":
    main()
