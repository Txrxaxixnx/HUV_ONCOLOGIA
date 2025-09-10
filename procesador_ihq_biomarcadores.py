#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extractor dedicado de biomarcadores IHQ (HER2, Ki-67, ER/PR, PD-L1, P16, Estudios Solicitados).
No modifica el esquema operativo estándar ni otros procesadores.
Genera un Excel extendido (55 + 8 columnas) solo para los PDFs IHQ seleccionados.
"""

import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

from ocr_processing import pdf_to_text_enhanced
import procesador_ihq as ihq


def _extract_biomarkers(text: str) -> dict:
    out = {
        "IHQ_HER2": "",
        "IHQ_KI-67": "",
        "IHQ_RECEPTOR_ESTROGENO": "",
        "IHQ_RECEPTOR_PROGESTAGENOS": "",
        "IHQ_PDL-1": "",
        "IHQ_ESTUDIOS_SOLICITADOS": "",
        "IHQ_P16_ESTADO": "",
        "IHQ_P16_PORCENTAJE": "",
    }

    # Estudios Solicitados
    m_est = re.search(r'(?i)estudios\s+solicitados:?\s*(?:se\s+realiz[oó]\s+tinci[oó]n\s+especial\s+para\s*)?([A-Z0-9 ,+/\.-]+)', text)
    if m_est:
        raw = m_est.group(1)
        tokens = []
        for token in re.split(r'[\s,;/]+', raw):
            t = token.strip().upper().replace(' ', '')
            if t:
                tokens.append(t)
        out["IHQ_ESTUDIOS_SOLICITADOS"] = ', '.join(dict.fromkeys(tokens))

    # P16
    m_p16_estado = re.search(r'(?i)\bP\s*16\b[^\n]*\b(positiv[oa]|negativ[oa])\b', text)
    if m_p16_estado:
        out["IHQ_P16_ESTADO"] = m_p16_estado.group(1).upper()
    elif re.search(r'(?i)\bP\s*16\b[^\n]*\b(en\s+bloque|difus[ao])\b', text):
        out["IHQ_P16_ESTADO"] = 'POSITIVO'
    m_p16_pct = re.search(r'(?i)\bP\s*16\b[^\n%]*?(\d{1,3})\s*%', text)
    if m_p16_pct:
        out["IHQ_P16_PORCENTAJE"] = m_p16_pct.group(1)

    # HER2 (score ± ISH/FISH)
    m_her2 = re.search(r'(?i)\bHER2(?:/NEU)?\b\s*[:\-]?\s*(0|1\+|2\+|3\+)', text)
    her2_score = m_her2.group(1) if m_her2 else ''
    m_ish = re.search(r'(?i)\b(?:ISH|FISH)\b[^\n]*\b(amplificad[oa]|no\s*amplificad[oa])\b', text)
    her2_ish = m_ish.group(1).upper().replace(' ', '_') if m_ish else ''
    out["IHQ_HER2"] = (f"{her2_score} ({'AMPLIFICADO' if 'AMPLIFICAD' in her2_ish else 'NO AMPLIFICADO'})".strip()
                       if her2_score and her2_ish else (her2_score or ('NO AMPLIFICADO' if her2_ish else '')))

    # Ki-67 (%)
    m_ki = re.search(r'(?i)\bK[IL]\s*[- ]?67\b[^\n%]*?(\d{1,3})\s*%', text)
    if m_ki:
        out["IHQ_KI-67"] = f"{m_ki.group(1)}%"

    # ER/RE
    m_re_estado = re.search(r'(?i)(receptor(?:\s+hormonal)?\s+de\s+estr(?:o|ó)geno|\bER\b)[^\n]*\b(positiv[oa]|negativ[oa])\b', text)
    m_re_pct = re.search(r'(?i)(receptor(?:\s+hormonal)?\s+de\s+estr(?:o|ó)geno|\bER\b)[^\n%]*?(\d{1,3})\s*%', text)
    re_txt = (m_re_estado.group(2).upper() if m_re_estado else '')
    re_pct = (m_re_pct.group(2) if m_re_pct and m_re_pct.lastindex and m_re_pct.lastindex>=2 else (m_re_pct.group(1) if m_re_pct else ''))
    out["IHQ_RECEPTOR_ESTROGENO"] = (f"{re_txt} {re_pct}%".strip() if re_txt and re_pct else (re_txt or (f"{re_pct}%" if re_pct else '')))

    # PR/RP
    m_rp_estado = re.search(r'(?i)(receptor(?:\s+hormonal)?\s+de\s+progest(?:erona|ágenos)|\bPR\b)[^\n]*\b(positiv[oa]|negativ[oa])\b', text)
    m_rp_pct = re.search(r'(?i)(receptor(?:\s+hormonal)?\s+de\s+progest(?:erona|ágenos)|\bPR\b)[^\n%]*?(\d{1,3})\s*%', text)
    rp_txt = (m_rp_estado.group(2).upper() if m_rp_estado else '')
    rp_pct = (m_rp_pct.group(2) if m_rp_pct and m_rp_pct.lastindex and m_rp_pct.lastindex>=2 else (m_rp_pct.group(1) if m_rp_pct else ''))
    out["IHQ_RECEPTOR_PROGESTAGENOS"] = (f"{rp_txt} {rp_pct}%".strip() if rp_txt and rp_pct else (rp_txt or (f"{rp_pct}%" if rp_pct else '')))

    # PD-L1 (TPS/CPS)
    m_tps = re.search(r'(?i)\bPD\s*-?L1\b[^\n]*\bTPS\b[^\n%]*?(\d{1,3})\s*%', text)
    m_cps = re.search(r'(?i)\bPD\s*-?L1\b[^\n]*\bCPS\b[^\n\d]*?(\d{1,3})', text)
    if m_tps and m_cps:
        out["IHQ_PDL-1"] = f"TPS {m_tps.group(1)}% (CPS {m_cps.group(1)})"
    elif m_tps:
        out["IHQ_PDL-1"] = f"TPS {m_tps.group(1)}%"
    elif m_cps:
        out["IHQ_PDL-1"] = f"CPS {m_cps.group(1)}"

    return out


def process_ihq_paths(pdf_paths: list[str], output_dir: str) -> Path:
    rows = []
    extended_columns = None

    for pdf in pdf_paths:
        text = pdf_to_text_enhanced(pdf)
        base = ihq.extract_ihq_data(text)
        base_rows = ihq.map_to_excel_format(base)
        if not base_rows:
            continue
        row = dict(base_rows[0])
        biomarkers = _extract_biomarkers(text)
        row.update(biomarkers)
        rows.append(row)

        if extended_columns is None:
            base_cols = list(base_rows[0].keys())
            extra_cols = list(biomarkers.keys())
            extended_columns = base_cols + extra_cols

    if not rows:
        raise RuntimeError("No se pudo extraer información de los PDFs IHQ seleccionados.")

    df = pd.DataFrame(rows)
    if extended_columns:
        df = df.reindex(columns=extended_columns)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = Path(output_dir) / f"Informe_IHQ_BIOMARCADORES_{timestamp}.xlsx"
    df.to_excel(output_path, index=False, engine='openpyxl')

    # Formato de encabezados
    wb = load_workbook(output_path)
    ws = wb.active
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
    for col in ws.columns:
        try:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws.column_dimensions[col[0].column_letter].width = min((max_length + 2), 60)
        except ValueError:
            pass
    wb.save(output_path)
    return output_path

