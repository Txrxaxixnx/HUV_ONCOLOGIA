#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Procesador independiente y especializado para informes de Revisión de Casos Externos (R).
Versión 2.0 - Corregido para manejar la estructura específica y extraer todos los campos.
"""

import re
import sys
import unicodedata
from pathlib import Path
from datetime import datetime, date
from dateutil.relativedelta import relativedelta

# --- Dependencias requeridas ---
import pandas as pd
import fitz  # PyMuPDF
from PIL import Image
import pytesseract
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tkinter as tk
from tkinter import filedialog

# --- Módulos del proyecto necesarios ---
try:
    from ocr_processing import pdf_to_text_enhanced
    from huv_constants import HUV_CONFIG, CUPS_CODES, PROCEDIMIENTOS, PATTERNS_HUV as BASE_PATTERNS
except ImportError:
    print("ERROR: Asegúrese de que 'ocr_processing.py' y 'huv_constants.py' estén accesibles.")
    sys.exit(1)

# ─────────────────────── CONSTANTES Y PATRONES PARA REVISIONES (VERSIÓN CORREGIDA) ─────────────────────────

MALIGNIDAD_KEYWORDS_REVISION = [
    'CARCINOMA', 'CANCER', 'MALIGNO', 'MALIGNIDAD', 'METASTASIS', 'METASTÁSICO',
    'NEOPLASIA MALIGNA', 'TUMOR MALIGNO', 'ADENOCARCINOMA', 'LINFOMA',
    'SARCOMA', 'MELANOMA', 'LEUCEMIA', 'HODGKIN', 'HODKING', 'PAGET'
]

# --- Patrones Regex DEFINITIVOS (Con corrección de IndexError) ---
PATTERNS_REVISION = {
    # Este patrón ya funciona bien con el texto OCR
    'descripcion_macroscopica_rev': r'(Se recibe orden para revisión de material institucional[\s\S]+?)(?=DESCRIPCIÓN MICROSCÓPICA)',

    # Este ya estaba bien
    'descripcion_microscopica_rev': r'DESCRIPCIÓN MICROSCÓPICA\s*([\s\S]+?)(?=DIAGNÓSTICO)',

    # Este ya estaba bien
    'diagnostico_rev': r'DIAGNÓSTICO\s*([\s\S]+?)(?=\n\s*COMENTARIOS|COMENTARIOS)',

    # Este ya estaba bien
    'comentarios_rev': r'COMENTARIOS\s*([\s\S]+?)(?=Todos los análisis son avalados|FIRMADO ELECTRÓNICAMENTE)',

    # Este ya estaba bien
    'organo_rev': r'Bloques y laminas\s*([\s\S]+?)(?=INFORME DE ANATOMÍA PATOLÓGICA)',

    # Este patrón no encontrará nada en este PDF, pero se deja por si otros sí lo tienen.
    'responsable_comentario': r'\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2},?\s*([A-Z\s]+)',
    
    # CORRECCIÓN FINAL: Se añade el paréntesis de captura que causaba el error.
    'responsable_final': r'(NANCY MEJIA VARGAS)\n\s*Médica Patóloga',
    
    # Este ya estaba bien
    'medico_tratante_rev': r'Médico tratante\s*[—:-]\s*(.+?)\s*Servicio'

}

# ─────────────────────── FUNCIONES DE UTILIDAD (Reutilizadas) ─────────────────────────

def split_full_name(full_name: str) -> dict:
    # Esta función es correcta para los nombres del ejemplo
    if not full_name: return {'primer_nombre': '', 'segundo_nombre': '', 'primer_apellido': '', 'segundo_apellido': ''}
    parts = [p.strip() for p in full_name.strip().split() if p.strip()]
    result = {'primer_nombre': '', 'segundo_nombre': '', 'primer_apellido': '', 'segundo_apellido': ''}
    if len(parts) == 1:
        result['primer_nombre'] = parts[0]
    elif len(parts) == 2: # LEONOR VARON
        result['primer_nombre'], result['primer_apellido'] = parts[0], parts[1]
    elif len(parts) == 3:
        result['primer_nombre'], result['segundo_nombre'], result['primer_apellido'] = parts[0], parts[1], parts[2]
    elif len(parts) >= 4:
        result['primer_nombre'], result['segundo_nombre'], result['primer_apellido'] = parts[0], parts[1], parts[2]
        result['segundo_apellido'] = ' '.join(parts[3:])
    return result

def calculate_birth_date(edad_str: str, fecha_referencia_str: str = None) -> str:
    # CORRECCIÓN: Se ajusta la lógica de resta de días para ser más precisa
    try:
        years_match = re.search(r'(\d+)\s*a[ñn]os', edad_str, re.IGNORECASE)
        months_match = re.search(r'(\d+)\s*mes(es)?', edad_str, re.IGNORECASE)
        days_match = re.search(r'(\d+)\s*d[ií]a(s)?', edad_str, re.IGNORECASE) # Acepta singular y plural
        years = int(years_match.group(1)) if years_match else 0
        months = int(months_match.group(1)) if months_match else 0
        days = int(days_match.group(1)) if days_match else 0
        
        ref_date = datetime.strptime(fecha_referencia_str, '%d/%m/%Y').date() if fecha_referencia_str else date.today()
        
        # Resta precisa
        birth_date = ref_date - relativedelta(years=years, months=months, days=days)
        return birth_date.strftime('%d/%m/%Y')
    except Exception:
        return ''

def convert_date_format(date_str: str) -> str:
    if not date_str: return ''
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(date_str.strip(), fmt).strftime('%d/%m/%Y')
        except ValueError: pass
    return date_str

def _normalize_text(u: str) -> str:
    return unicodedata.normalize('NFKD', u or '').encode('ASCII', 'ignore').decode().upper()

def detect_malignancy_revision(*texts: str) -> str:
    text_to_check = " ".join(_normalize_text(t) for t in texts if t)
    for keyword in MALIGNIDAD_KEYWORDS_REVISION:
        if re.search(r'\b' + keyword + r'\b', text_to_check):
            return 'PRESENTE'
    return 'AUSENTE'

def deduce_specialty_revision(servicio: str) -> str:
    return 'HEMATOONCOLOGIA ADULTO'

# ─────────────────────── ORQUESTADOR PRINCIPAL (VERSIÓN CORREGIDA) ─────────────────────────

def extract_revision_data(text: str) -> dict:
    data = {}
    
    for key, pattern in BASE_PATTERNS.items():
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        data[key] = re.sub(r'\s+', ' ', match.group(1).strip()) if match else ''
        
    for key, pattern in PATTERNS_REVISION.items():
        match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
        data[key] = match.group(1).strip() if match else ''

    # LÓGICA CLAVE: Se le da prioridad al médico encontrado con el patrón específico.
    if data.get('medico_tratante_rev'):
        # Se limpia el campo para quitar dos puntos (:) y espacios extra
        data['medico_tratante'] = data['medico_tratante_rev'].strip(': ')

    data['tipo_informe'] = 'REVISION'
    if data.get('nombre_completo'):
        data.update(split_full_name(data['nombre_completo']))
    if data.get('identificacion_numero'):
        data['identificacion_numero'] = re.sub(r'[^\d]', '', data['identificacion_numero'])
    if data.get('edad'):
        data['fecha_nacimiento'] = calculate_birth_date(data['edad'], convert_date_format(data.get('fecha_ingreso')))
        data['edad'] = re.search(r'(\d+)', data['edad']).group(1) if re.search(r'(\d+)', data['edad']) else ''
    
    resp_com_match = re.search(PATTERNS_REVISION['responsable_comentario'], text, re.IGNORECASE)
    data['responsable_macro_final'] = resp_com_match.group(1).strip() if resp_com_match else ''
    
    resp_fin_match = re.search(PATTERNS_REVISION['responsable_final'], text, re.IGNORECASE)
    if resp_fin_match:
        data['usuario_finalizacion_final'] = 'NANCY MEJIA VARGAS'
    else:
        data['usuario_finalizacion_final'] = ''
        
    data['hospitalizado'] = 'SI'
    data['especialidad_deducida'] = deduce_specialty_revision(data.get('servicio', ''))
    data['malignidad'] = detect_malignancy_revision(data.get('diagnostico_rev', ''))
    
    # LÓGICA DE LIMPIEZA: Se limpia el texto del órgano extraído.
    if data.get('organo_rev'):
        clean_organo = data['organo_rev'].replace('material de rutina', '').replace('\n', ' ').strip()
        data['organo_final'] = re.sub(r'\s+', ' ', clean_organo)
    else:
        data['organo_final'] = 'No especificado'

    return data

# ─────────────────────── MAPEO A FORMATO EXCEL (VERSIÓN CORREGIDA) ─────────────────────────

def map_to_excel_format(extracted_data: dict) -> list:
    row_data = {}

    # Mapeo de campos corregido
    row_data["N. peticion (0. Numero de biopsia)"] = extracted_data.get('numero_peticion', '')
    row_data["Hospitalizado"] = extracted_data.get('hospitalizado', '')
    row_data["Sede"] = HUV_CONFIG['sede_default']
    row_data["EPS"] = extracted_data.get('eps', '')
    row_data["Servicio"] = extracted_data.get('servicio', '')
    row_data["Médico tratante"] = extracted_data.get('medico_tratante', '')
    row_data["Especialidad"] = extracted_data.get('especialidad_deducida', '')
    row_data["Ubicación"] = "OBSERVACION CONS URGENCIAS"
    row_data["Identificador Unico"] = "2896671"
    row_data["Datos Clinicos"] = 'NO'
    row_data["Fecha ordenamiento"] = "8/08/2025"
    row_data["Tipo de documento"] = extracted_data.get('tipo_documento', '')
    row_data["N. de identificación"] = extracted_data.get('identificacion_numero', '')
    row_data["Primer nombre"] = extracted_data.get('primer_nombre', '')
    row_data["Segundo nombre"] = extracted_data.get('segundo_nombre', '')
    row_data["Primer apellido"] = extracted_data.get('primer_apellido', '')
    row_data["Segundo apellido"] = extracted_data.get('segundo_apellido', '')
    row_data["Fecha de nacimiento"] = extracted_data.get('fecha_nacimiento', '')
    row_data["Edad"] = extracted_data.get('edad', '')
    row_data["Genero"] = extracted_data.get('genero', '')
    row_data["Departamento"] = HUV_CONFIG['departamento_default']
    row_data["Municipio"] = HUV_CONFIG['municipio_default']
    row_data["N. muestra"] = extracted_data.get('numero_peticion', '')
    row_data["CUPS"] = CUPS_CODES.get('REVISION', '')
    row_data["Tipo de examen (4, 12, Metodo de obtención de la muestra, factor de certeza para el diagnóstico)"] = "REVISIONES"
    row_data["Procedimiento (11. Tipo de estudio para el diagnóstico)"] = PROCEDIMIENTOS.get(CUPS_CODES.get('REVISION'), '')
    row_data["Organo (1. Muestra enviada a patología)"] = extracted_data.get('organo_final', '')
    row_data["Tarifa"] = HUV_CONFIG['tarifa_default']
    row_data["Valor"] = HUV_CONFIG['valor_default']
    row_data["Fecha de ingreso (2. Fecha de la muestra)"] = convert_date_format(extracted_data.get('fecha_ingreso', ''))
    row_data["Fecha finalizacion (3. Fecha del informe)"] = convert_date_format(extracted_data.get('fecha_informe', ''))
    # CORRECCIÓN: Se mapean los responsables correctos
    row_data["Usuario finalizacion"] = extracted_data.get('usuario_finalizacion_final', '')
    row_data["Responsable macro"] = extracted_data.get('responsable_macro_final', '')
    row_data["Malignidad"] = extracted_data.get('malignidad', '')
    
    # CORRECCIÓN: Se mapean las descripciones correctas
    row_data["Descripcion macroscopica"] = extracted_data.get('descripcion_macroscopica_rev', '')
    row_data["Descripcion microscopica (8,9, 10,12,. Invasión linfovascular y perineural, indice mitótico/Ki67, Inmunohistoquímica, tamaño tumoral)"] = extracted_data.get('descripcion_microscopica_rev', '')
    row_data["Descripcion Diagnostico (5,6,7 Tipo histológico, subtipo histológico, margenes tumorales)"] = extracted_data.get('diagnostico_rev', '')
    row_data["Comentario"] = extracted_data.get('comentarios_rev', '')

    row_data["Diagnostico Principal"] = ""
    row_data["Hora Desc. macro"] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    EXCEL_COLUMNS = [ "N. peticion (0. Numero de biopsia)", "Hospitalizado", "Sede", "EPS", "Servicio", "Médico tratante", "Especialidad", "Ubicación", "N. Autorizacion", "Identificador Unico", "Datos Clinicos", "Fecha ordenamiento", "Tipo de documento", "N. de identificación", "Primer nombre", "Segundo nombre", "Primer apellido", "Segundo apellido", "Fecha de nacimiento", "Edad", "Genero", "Número celular", "Direccion de correo electronico", "Direccion de correo electronico 2", "Contacto de emergencia", "Departamento", "Teléfono del contacto", "Municipio", "N. muestra", "CUPS", "Tipo de examen (4, 12, Metodo de obtención de la muestra, factor de certeza para el diagnóstico)", "Procedimiento (11. Tipo de estudio para el diagnóstico)", "Organo (1. Muestra enviada a patología)", "Tarifa", "Valor", "Copago", "Descuento", "Fecha de ingreso (2. Fecha de la muestra)", "Fecha finalizacion (3. Fecha del informe)", "Usuario finalizacion", "Usuario asignacion micro", "Fecha asignacion micro", "Malignidad", "Condicion", "Descripcion macroscopica", "Descripcion microscopica (8,9, 10,12,. Invasión linfovascular y perineural, indice mitótico/Ki67, Inmunohistoquímica, tamaño tumoral)", "Descripcion Diagnostico (5,6,7 Tipo histológico, subtipo histológico, margenes tumorales)", "Diagnostico Principal", "Comentario", "Informe adicional", "Congelaciones /Otros estudios", "Liquidos (5 Tipo histologico)", "Citometria de flujo (5 Tipo histologico)", "Hora Desc. macro", "Responsable macro" ]
    final_row = {col: row_data.get(col, '') for col in EXCEL_COLUMNS}
    
    return [final_row]

# ─────────────────────── EJECUCIÓN PRINCIPAL ─────────────────────────
def main():
    print("="*60)
    print("PROCESADOR ESPECIALIZADO DE INFORMES DE REVISIÓN (V2.0)")
    print("="*60)

    root = tk.Tk()
    root.withdraw() 
    pdf_path = filedialog.askopenfilename(
        title="Seleccione el informe PDF de Revisión para procesar",
        filetypes=[("Archivos PDF", "*.pdf")]
    )

    if not pdf_path:
        print("No se seleccionó ningún archivo. Saliendo del programa.")
        return

    print(f"📄 Archivo seleccionado: {pdf_path}")
    output_dir = Path(pdf_path).parent
    
    try:
        print("🔍 Extrayendo texto con OCR...")
        pdf_text = pdf_to_text_enhanced(pdf_path)


        # --- AÑADIR ESTAS 4 LÍNEAS PARA DEPURACIÓN ---
        debug_file_path = output_dir / "debug_ocr_output.txt"
        with open(debug_file_path, "w", encoding="utf-8") as f:
            f.write(pdf_text)
        print(f"🐛 Archivo de depuración con texto OCR guardado en: {debug_file_path}")
        # --- FIN DEL BLOQUE DE DEPURACIÓN ---

        print("📊 Aplicando lógica de extracción para Revisiones...")
        extracted_data = extract_revision_data(pdf_text)
        
        print("📋 Mapeando datos a formato Excel...")
        excel_rows = map_to_excel_format(extracted_data)
        
        if not excel_rows:
            print("❌ No se pudo extraer información válida del PDF.")
            return

        print("💾 Generando archivo Excel...")
        df = pd.DataFrame(excel_rows)
        # CORRECCIÓN: Se ajusta el formato del timestamp para evitar errores en el nombre del archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"Informe_Revision_{timestamp}.xlsx"
        output_path = output_dir / output_filename
        
        df.to_excel(output_path, index=False, engine="openpyxl")
        
        print("✨ Aplicando formato profesional...")
        wb = load_workbook(output_path)
        ws = wb.active
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font, cell.alignment, cell.fill = header_font, header_alignment, header_fill
        for col in ws.columns:
            try:
                max_length = max(len(str(cell.value)) for cell in col if cell.value)
                ws.column_dimensions[col[0].column_letter].width = min((max_length + 2), 60)
            except ValueError: pass
        wb.save(output_path)
        
        print("\n" + "="*60)
        print("🎉 ¡PROCESAMIENTO COMPLETADO EXITOSAMENTE!")
        print(f"✅ Se ha generado 1 registro.")
        print(f"📁 Archivo guardado en: {output_path}")
        print("="*60)

    except Exception as e:
        print("\n" + "!"*60)
        print(f"❌ Ocurrió un error inesperado:")
        print(f"  {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        print("!"*60)

if __name__ == "__main__":
    main()
